import tkinter as tk
from tkinter import ttk, StringVar, messagebox
import tkinter.font
import guiFiles.guihelperfunc as make
import guiFiles.guivalidatefunc as validate
import guiFiles.initinstruments as init
import guiFiles.auxfunclib as aux
from guiFiles.tk_tooltip import CreateToolTip
from guiFiles.configmgr import ConfigMgr
from guiFiles.killthread import KillableThread
from guiFiles.runreportgen import RunReportGen
import pverifyDrivers.purescope as purescope

import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.animation as animation
from matplotlib import style
style.use('ggplot')

from time import sleep
import json
import os
import datetime as dt
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Alignment

DEBUG = 1
LOADTIMEOFF = 10 #Time off between load steps on snapshot option

class EffFrame(tk.Frame):

    tempList = None
    currList = None
    currList2 = None
    freqList = None
    pvinList = None
    voutList = None
    vccList = None
    modeList = None
    enList = None
    biasList = None
    delayList = None

    load = None
    keith = None
    vcc = None
    vin = None
    scope = None
    snapscope = None
    dongle = None
    en = None

    def __init__(self, parent):
        super(EffFrame, self).__init__(parent)
        self.run = RunReportGen()
        self.config(background='white')
        self.coolDownTemp = 34
        self.r = 0
        self.c = 0
        self.style = ttk.Style()
        # self.style.theme_use('default')
        self.style.configure(
            'Treeview',
            background='#D3D3D3',
            foreground='black',
            rowheight=25,
            fieldbackground='#D3D3D3'
        )
        self.style.map(
            'Treeview',
            background=[('selected', '#347083')],
        )

        # region JIRA
        self.jira = tk.Frame(self, background='white')
        self.jira.grid(row=self.r, column=self.c, padx=10, sticky='w')
        self.build_box_jira()
        
        self.r += 1
        #endregion

        # region Chamber Temperatue
        self.Temp = tk.LabelFrame(self, text='Temperature Steps (\u00b0C)', background='white')
        self.Temp.grid(row=self.r, column=self.c, padx=2, sticky='w')
        self.build_box_temp()
        self.c += 1
        # endregion

        # region Status Bar
        self.statusBar = tk.LabelFrame(self, text='Status', background='white')
        self.statusBar.grid(row=self.r, column=self.c, padx=10, sticky='nsew')
        self.statusLabel = tk.Label(
            self.statusBar,
            text='WELCOME!',
            anchor=tk.CENTER,
            foreground='#0FFF50',
            background='black',
            font=('Arial Bold', 12)
        )
        self.statusLabel.pack(expand=1, fill='both')
        self.r += 1
        self.c -= 1
        # endregion

        # Setup box
        self.setupLblFrame = tk.LabelFrame(self, text='Setup', background='#A4E1FF')
        self.setupLblFrame.grid(row=self.r, column=self.c, padx=2, pady=5, sticky='nw')
        self.setup_box()
        self.c += 1

        # Graph
        self.graphFrame = ttk.Frame(self)
        self.graphFrame.grid(row=self.r, column=self.c, rowspan=30, padx=2, pady=5, sticky='nw')
        self.xAxis = list()
        self.yAxis = list()
        self.f = Figure(figsize=(5, 5), dpi=100)
        self.a = self.f.add_subplot(111)
        self.a.set_ylim(70, 100)
        self.graphF = ttk.Frame(self.graphFrame)
        self.graphF.grid(row=0, column=0)
        self.plotCanvas = FigureCanvasTkAgg(self.f, master=self.graphF)
        self.plotCanvas.get_tk_widget().grid(row=0, column=0, padx=2, pady=2)
        # self.graph_box()
        self.r += 1
        self.c -= 1

        # Table
        self.tableFrame = tk.Frame(self)
        self.tableFrame.grid(row=self.r, column=self.c, padx=2, rowspan=28, sticky='nsew')
        self.tableFrame.grid_propagate(0)
        tableScrollY = tk.Scrollbar(self.tableFrame)
        tableScrollY.pack(side=tk.RIGHT, fill=tk.Y)
        tableScrollX = tk.Scrollbar(self.tableFrame, orient='horizontal')
        tableScrollX.pack(side=tk.BOTTOM, fill=tk.X)
        self.table = ttk.Treeview(
            self.tableFrame,
            yscrollcommand=tableScrollY.set,
            xscrollcommand=tableScrollX.set,
            selectmode='extended',
        )
        self.table.pack(fill=tk.BOTH, expand=True)
        tableScrollY.config(command=self.table.yview)
        tableScrollX.config(command=self.table.xview)
        self.table_box()
        self.r += 28

        # region BUTTONS
        BOLDFONT = tk.font.Font(weight='bold')
        btnFrame = ttk.Frame(self)
        btnFrame.grid(row=self.r, column=self.c, sticky='se', padx=5)
        # region KILL TEST BUTTON
        self.killBtn = tk.Button(
            btnFrame,
            text='KILL TEST',
            padx=5,
            font=BOLDFONT,
            background='red',
            foreground='white',
            command=self.kill_test,
        )
        self.killBtn.grid(row=0, column=0, padx=5)
        CreateToolTip(self.killBtn, 'Most or ALL DATA will be LOST!')
        self.killBtn['state'] = 'disabled'
        # endregion

        # region RUN BUTTON
        self.runButton = tk.Button(
            btnFrame,
            text='Run Test',
            padx=5,
            font=BOLDFONT,
            background='#003833',
            foreground='white',
            command=self.run_tests,
        )
        self.runButton.grid(row=0, column=1, padx=5)
        # endregion
        # endregion

    def build_box_temp(self):
        self.tempFrame = tk.Frame(self.Temp, background='#FFB4A4')
        self.tempFrame.grid(row=3, column=0, columnspan=5, rowspan=3)
        self.varTempStep1 = StringVar()
        self.varTempStep1.set(ConfigMgr.tempSteps['1'])
        self.varTempStep2 = StringVar()
        self.varTempStep2.set(ConfigMgr.tempSteps['2'])
        self.varTempStep3 = StringVar()
        self.varTempStep3.set(ConfigMgr.tempSteps['3'])
        self.varTempStep4 = StringVar()
        self.varTempStep4.set(ConfigMgr.tempSteps['4'])
        self.varTempStep5 = StringVar()
        self.varTempStep5.set(ConfigMgr.tempSteps['5'])
        self.varTempStep6 = StringVar()
        self.varTempStep6.set(ConfigMgr.tempSteps['6'])
        tempArgList = [
            self.tempFrame,
            ConfigMgr.tempSteps,
            validate.temperature,
            self.varTempStep1,
            self.varTempStep2,
            self.varTempStep3,
            self.varTempStep4,
            self.varTempStep5,
            self.varTempStep6,
        ]
        make.temperature_box(tempArgList)

    def build_box_jira(self):
        self.jiraFrame = tk.Frame(self.jira, background='#FFB4A4')
        self.jiraFrame.grid(row=3, column=0, columnspan=5, rowspan=3)
        make.jira_box()

    def setup_box(self):
        # region EN
        enLblFrame = tk.LabelFrame(self.setupLblFrame, text='Enable (V)', background='#A4E1FF')
        enLblFrame.grid(row=0, column=0, rowspan=3, padx=2, pady=5, sticky='nsw')
        self.enRiseTimeFrame101 = tk.Frame(self.setupLblFrame)
        self.varEnOpt101 = StringVar()
        self.varEnOpt101.set(ConfigMgr.testConditions101['enOpt'])
        self.varEn1101 = StringVar()
        self.varEn1101.set(ConfigMgr.testConditions101['en1'])
        self.varEn2101 = StringVar()
        self.varEn2101.set(ConfigMgr.testConditions101['en2'])
        self.varEn3101 = StringVar()
        self.varEn3101.set(ConfigMgr.testConditions101['en3'])
        self.varEnPvin = StringVar()
        self.varEnPvin.set(ConfigMgr.testConditions101['enPvin'])
        self.enBoxArgList101 = [
            enLblFrame,
            self.varEnOpt101,
            ConfigMgr.testConditions101,
            self.varEn1101,
            self.varEn2101,
            self.varEn3101,
            validate.en_ext,
            validate.en_entry,
            self.enRiseTimeFrame101,
            self.varEnPvin,
        ]
        make.en_box(self.enBoxArgList101)
        validate.en_ext(self.enBoxArgList101)
        # endregion

        #region Vcc Steps
        vccFrame = tk.LabelFrame(self.setupLblFrame, text='Vcc Steps (V)', background='#A4E1FF')
        vccFrame.grid(row=0, column=1, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartVcc = StringVar()
        lbl1 = tk.Label(vccFrame, text='Start Vcc:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            vccFrame,
            textvariable=self.varStartVcc,
            validate='focusout',
            validatecommand=self.validate_vcc,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartVcc.set(ConfigMgr.testConditions101['startVcc'])
        self.varEndVcc = StringVar()
        lbl2 = tk.Label(vccFrame, text='End Vcc:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            vccFrame,
            textvariable=self.varEndVcc,
            validate='focusout',
            validatecommand=self.validate_vcc,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndVcc.set(ConfigMgr.testConditions101['endVcc'])
        self.varStepVcc = StringVar()
        lbl3 = tk.Label(vccFrame, text='Step Vcc:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            vccFrame,
            textvariable=self.varStepVcc,
            validate='focusout',
            validatecommand=self.validate_vcc,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepVcc.set(ConfigMgr.testConditions101['stepVcc'])
        #endregion

        #region Vin Steps
        vinFrame = tk.LabelFrame(self.setupLblFrame, text='Vin Steps (V)', background='#A4E1FF')
        vinFrame.grid(row=0, column=2, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartVin = StringVar()
        self.varEndVin = StringVar()
        self.varStepVin = StringVar()
        self.varVin01 = StringVar()
        self.varVin02 = StringVar()
        self.varVin03 = StringVar()
        self.varVin04 = StringVar()
        self.varVin05 = StringVar()
        self.varVin06 = StringVar()
        self.varVinOpt = StringVar()
        self.varStartVin.set(ConfigMgr.testConditions101['startPVin'])
        self.varEndVin.set(ConfigMgr.testConditions101['endPVin'])
        self.varStepVin.set(ConfigMgr.testConditions101['stepPVin'])
        self.varVin01.set(ConfigMgr.testConditions101['PVin1'])
        self.varVin02.set(ConfigMgr.testConditions101['PVin2'])
        self.varVin03.set(ConfigMgr.testConditions101['PVin3'])
        self.varVin04.set(ConfigMgr.testConditions101['PVin4'])
        self.varVin05.set(ConfigMgr.testConditions101['PVin5'])
        self.varVin06.set(ConfigMgr.testConditions101['PVin6'])
        self.varVinOpt.set(ConfigMgr.testConditions101['pvinOpt'])

        #region incremental frame
        self.incrFrame = tk.Frame(vinFrame, background='#A4E1FF')
        self.incrFrame.grid(row=1, column=0)
        lbl1 = tk.Label(self.incrFrame, text='Start Vin:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            self.incrFrame,
            textvariable=self.varStartVin,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        lbl2 = tk.Label(self.incrFrame, text='End Vin:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            self.incrFrame,
            textvariable=self.varEndVin,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        lbl3 = tk.Label(self.incrFrame, text='Step Vin:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            self.incrFrame,
            textvariable=self.varStepVin,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        #endregion

        #region fixed steps frame
        self.fixedFrame = tk.Frame(vinFrame, background='#A4E1FF')
        self.fixedFrame.grid(row=1, column=0)
        vin1Label = ttk.Label(self.fixedFrame, text='1:', background='#A4E1FF')
        vin1Label.grid(row=0, column=0, padx=2, sticky='e')
        vin1Entry = ttk.Entry(
            self.fixedFrame,
            textvariable=self.varVin01,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        vin1Entry.grid(row=0, column=1, pady=5, sticky='w')
        vin2Label = ttk.Label(self.fixedFrame, text='2:', background='#A4E1FF')
        vin2Label.grid(row=0, column=2, pady=5, padx=2, sticky='e')
        vin2Entry = ttk.Entry(
            self.fixedFrame,
            textvariable=self.varVin02,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        vin2Entry.grid(row=0, column=3, pady=5, sticky='w')
        vin3Label = ttk.Label(self.fixedFrame, text='3:', background='#A4E1FF')
        vin3Label.grid(row=1, column=0, pady=5, padx=2, sticky='e')
        vin3Entry = ttk.Entry(
            self.fixedFrame,
            textvariable=self.varVin03,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        vin3Entry.grid(row=1, column=1, pady=5, sticky='w')
        vin4Label = ttk.Label(self.fixedFrame, text='4:', background='#A4E1FF')
        vin4Label.grid(row=1, column=2, pady=5, padx=2, sticky='e')
        vin4Entry = ttk.Entry(
            self.fixedFrame,
            textvariable=self.varVin04,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        vin4Entry.grid(row=1, column=3, pady=5, sticky='w')
        vin5Label = ttk.Label(self.fixedFrame, text='5:', background='#A4E1FF')
        vin5Label.grid(row=2, column=0, pady=5, padx=2, sticky='e')
        vin5Entry = ttk.Entry(
            self.fixedFrame,
            textvariable=self.varVin05,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        vin5Entry.grid(row=2, column=1, pady=5, sticky='w')
        vin6Label = ttk.Label(self.fixedFrame, text='6:', background='#A4E1FF')
        vin6Label.grid(row=2, column=2, pady=5, padx=2, sticky='e')
        vin6Entry = ttk.Entry(
            self.fixedFrame,
            textvariable=self.varVin06,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        vin6Entry.grid(row=2, column=3, pady=5, sticky='w')
        #endregion

        #region vin option
        vinOptFrame = tk.Frame(vinFrame, background='#A4E1FF')
        vinOptFrame.grid(row=0, column=0)
        vinOptRadioButton1 = tk.Radiobutton(
            vinOptFrame,
            text='Fixed Steps',
            variable=self.varVinOpt,
            value='fixed',
            command=self.validate_vin_opt,
            background='#A4E1FF',
        )
        vinOptRadioButton2 = tk.Radiobutton(
            vinOptFrame,
            text='Increment Steps',
            variable=self.varVinOpt,
            value='incr',
            command=self.validate_vin_opt,
            background='#A4E1FF'
        )
        vinOptRadioButton1.grid(row=1, column=0, pady=2, sticky='w')
        vinOptRadioButton2.grid(row=2, column=0, pady=2, sticky='w')
        self.validate_vin_opt()
        #endregion

        #endregion

        #region Frequency Steps
        fswFrame = tk.LabelFrame(self.setupLblFrame, text='Frequency Steps (kHz)', background='#A4E1FF')
        fswFrame.grid(row=0, column=3, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartFsw = StringVar()
        self.varEndFsw = StringVar()
        self.varStepFsw = StringVar()
        self.varFsw01 = StringVar()
        self.varFsw02 = StringVar()
        self.varFsw03 = StringVar()
        self.varFsw04 = StringVar()
        self.varFsw05 = StringVar()
        self.varFsw06 = StringVar()
        self.varFswOpt = StringVar()
        self.varStartFsw.set(ConfigMgr.testConditions101['startFsw'])
        self.varEndFsw.set(ConfigMgr.testConditions101['endFsw'])
        self.varStepFsw.set(ConfigMgr.testConditions101['stepFsw'])
        self.varFsw01.set(ConfigMgr.testConditions101['fsw1'])
        self.varFsw02.set(ConfigMgr.testConditions101['fsw2'])
        self.varFsw03.set(ConfigMgr.testConditions101['fsw3'])
        self.varFsw04.set(ConfigMgr.testConditions101['fsw4'])
        self.varFsw05.set(ConfigMgr.testConditions101['fsw5'])
        self.varFsw06.set(ConfigMgr.testConditions101['fsw6'])
        self.varFswOpt.set(ConfigMgr.testConditions101['fswOpt'])

        #region Incremental Frequencies
        self.incrFreqFrame = tk.Frame(fswFrame, background='#A4E1FF')
        self.incrFreqFrame.grid(row=1, column=0)
        lbl1 = tk.Label(self.incrFreqFrame, text='Start Frequency:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            self.incrFreqFrame,
            textvariable=self.varStartFsw,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        lbl2 = tk.Label(self.incrFreqFrame, text='End Frequency:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            self.incrFreqFrame,
            textvariable=self.varEndFsw,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        lbl3 = tk.Label(self.incrFreqFrame, text='Step Frequency:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            self.incrFreqFrame,
            textvariable=self.varStepFsw,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        #endregion

        #region Fixed Steps Frequency
        self.fswFixedFrame = tk.Frame(fswFrame, background='#A4E1FF')
        self.fswFixedFrame.grid(row=1, column=0)
        fsw1Label = ttk.Label(self.fswFixedFrame, text='1:', background='#A4E1FF')
        fsw1Label.grid(row=0, column=0, padx=2, sticky='e')
        fsw1Entry = ttk.Entry(
            self.fswFixedFrame,
            textvariable=self.varFsw01,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        fsw1Entry.grid(row=0, column=1, pady=5, sticky='w')
        fsw2Label = ttk.Label(self.fswFixedFrame, text='2:', background='#A4E1FF')
        fsw2Label.grid(row=0, column=2, pady=5, padx=2, sticky='e')
        fsw2Entry = ttk.Entry(
            self.fswFixedFrame,
            textvariable=self.varFsw02,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        fsw2Entry.grid(row=0, column=3, pady=5, sticky='w')
        fsw3Label = ttk.Label(self.fswFixedFrame, text='3:', background='#A4E1FF')
        fsw3Label.grid(row=1, column=0, pady=5, padx=2, sticky='e')
        fsw3Entry = ttk.Entry(
            self.fswFixedFrame,
            textvariable=self.varFsw03,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        fsw3Entry.grid(row=1, column=1, pady=5, sticky='w')
        fsw4Label = ttk.Label(self.fswFixedFrame, text='4:', background='#A4E1FF')
        fsw4Label.grid(row=1, column=2, pady=5, padx=2, sticky='e')
        fsw4Entry = ttk.Entry(
            self.fswFixedFrame,
            textvariable=self.varFsw04,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        fsw4Entry.grid(row=1, column=3, pady=5, sticky='w')
        fsw5Label = ttk.Label(self.fswFixedFrame, text='5:', background='#A4E1FF')
        fsw5Label.grid(row=2, column=0, pady=5, padx=2, sticky='e')
        fsw5Entry = ttk.Entry(
            self.fswFixedFrame,
            textvariable=self.varFsw05,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        fsw5Entry.grid(row=2, column=1, pady=5, sticky='w')
        fsw6Label = ttk.Label(self.fswFixedFrame, text='6:', background='#A4E1FF')
        fsw6Label.grid(row=2, column=2, pady=5, padx=2, sticky='e')
        fsw6Entry = ttk.Entry(
            self.fswFixedFrame,
            textvariable=self.varFsw06,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        fsw6Entry.grid(row=2, column=3, pady=5, sticky='w')
        #endregion

        #region Frequency Option
        fswOptFrame = tk.Frame(fswFrame, background='#A4E1FF')
        fswOptFrame.grid(row=0, column=0)
        fswOptRadioButton1 = tk.Radiobutton(
            fswOptFrame,
            text='Fixed Steps',
            variable=self.varFswOpt,
            value='fixed',
            command=self.validate_fsw_opt,
            background='#A4E1FF',
        )
        fswOptRadioButton2 = tk.Radiobutton(
            fswOptFrame,
            text='Increment Steps',
            variable=self.varFswOpt,
            value='incr',
            command=self.validate_fsw_opt,
            background='#A4E1FF'
        )
        fswOptRadioButton1.grid(row=1, column=0, pady=2, sticky='w')
        fswOptRadioButton2.grid(row=2, column=0, pady=2, sticky='w')
        self.validate_fsw_opt()
        #endregion

        #endregion

        #region Load Steps
        loadFrame = tk.LabelFrame(self.setupLblFrame, text='Load Steps (A)', background='#A4E1FF')
        loadFrame.grid(row=0, column=4, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartCurrent = StringVar()
        lbl1 = tk.Label(loadFrame, text='Start Current:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            loadFrame,
            textvariable=self.varStartCurrent,
            validate='focusout',
            validatecommand=self.validate_current,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartCurrent.set(ConfigMgr.testConditions101['startCurr'])
        self.varEndCurrent = StringVar()
        lbl2 = tk.Label(loadFrame, text='End Current:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            loadFrame,
            textvariable=self.varEndCurrent,
            validate='focusout',
            validatecommand=self.validate_current,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndCurrent.set(ConfigMgr.testConditions101['endCurr'])
        self.varStepCurrent = StringVar()
        lbl3 = tk.Label(loadFrame, text='Step Current:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            loadFrame,
            textvariable=self.varStepCurrent,
            validate='focusout',
            validatecommand=self.validate_current,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepCurrent.set(ConfigMgr.testConditions101['stepCurr'])
        #endregion

        #region Vout Steps
        voutFrame = tk.LabelFrame(self.setupLblFrame, text='Vout Steps (V)', background='#A4E1FF')
        voutFrame.grid(row=0, column=5, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartVout = StringVar()
        self.varEndVout = StringVar()
        self.varStepVout = StringVar()
        self.varVout01 = StringVar()
        self.varVout02 = StringVar()
        self.varVout03 = StringVar()
        self.varVout04 = StringVar()
        self.varVout05 = StringVar()
        self.varVout06 = StringVar()
        self.varVoutOpt = StringVar()
        self.varStartVout.set(ConfigMgr.testConditions101['startVout'])
        self.varEndVout.set(ConfigMgr.testConditions101['endVout'])
        self.varStepVout.set(ConfigMgr.testConditions101['stepVout'])
        self.varVout01.set(ConfigMgr.testConditions101['vout1'])
        self.varVout02.set(ConfigMgr.testConditions101['vout2'])
        self.varVout03.set(ConfigMgr.testConditions101['vout3'])
        self.varVout04.set(ConfigMgr.testConditions101['vout4'])
        self.varVout05.set(ConfigMgr.testConditions101['vout5'])
        self.varVout06.set(ConfigMgr.testConditions101['vout6'])
        self.varVoutOpt.set(ConfigMgr.testConditions101['voutOpt'])

        #region Incremental Vout
        self.incrVoutFrame = tk.Frame(voutFrame, background='#A4E1FF')
        self.incrVoutFrame.grid(row=1, column=0)
        lbl1 = tk.Label(self.incrVoutFrame, text='Start Vout:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            self.incrVoutFrame,
            textvariable=self.varStartVout,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        lbl2 = tk.Label(self.incrVoutFrame, text='End Vout:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            self.incrVoutFrame,
            textvariable=self.varEndVout,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        lbl3 = tk.Label(self.incrVoutFrame, text='Step Vout:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            self.incrVoutFrame,
            textvariable=self.varStepVout,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        #endregion

        #region Fixed Vout
        self.fixedVoutFrame = tk.Frame(voutFrame, background='#A4E1FF')
        self.fixedVoutFrame.grid(row=1, column=0)
        vout1Label = ttk.Label(self.fixedVoutFrame, text='1:', background='#A4E1FF')
        vout1Label.grid(row=0, column=0, padx=2, sticky='e')
        vout1Entry = ttk.Entry(
            self.fixedVoutFrame,
            textvariable=self.varVout01,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        vout1Entry.grid(row=0, column=1, pady=5, sticky='w')
        vout2Label = ttk.Label(self.fixedVoutFrame, text='2:', background='#A4E1FF')
        vout2Label.grid(row=0, column=2, pady=5, padx=2, sticky='e')
        vout2Entry = ttk.Entry(
            self.fixedVoutFrame,
            textvariable=self.varVout02,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        vout2Entry.grid(row=0, column=3, pady=5, sticky='w')
        vout3Label = ttk.Label(self.fixedVoutFrame, text='3:', background='#A4E1FF')
        vout3Label.grid(row=1, column=0, pady=5, padx=2, sticky='e')
        vout3Entry = ttk.Entry(
            self.fixedVoutFrame,
            textvariable=self.varVout03,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        vout3Entry.grid(row=1, column=1, pady=5, sticky='w')
        vout4Label = ttk.Label(self.fixedVoutFrame, text='4:', background='#A4E1FF')
        vout4Label.grid(row=1, column=2, pady=5, padx=2, sticky='e')
        vout4Entry = ttk.Entry(
            self.fixedVoutFrame,
            textvariable=self.varVout04,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        vout4Entry.grid(row=1, column=3, pady=5, sticky='w')
        vout5Label = ttk.Label(self.fixedVoutFrame, text='5:', background='#A4E1FF')
        vout5Label.grid(row=2, column=0, pady=5, padx=2, sticky='e')
        vout5Entry = ttk.Entry(
            self.fixedVoutFrame,
            textvariable=self.varVout05,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        vout5Entry.grid(row=2, column=1, pady=5, sticky='w')
        vout6Label = ttk.Label(self.fixedVoutFrame, text='6:', background='#A4E1FF')
        vout6Label.grid(row=2, column=2, pady=5, padx=2, sticky='e')
        vout6Entry = ttk.Entry(
            self.fixedVoutFrame,
            textvariable=self.varVout06,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        vout6Entry.grid(row=2, column=3, pady=5, sticky='w')
        #endregion

        #region Vout Option
        voutOptFrame = tk.Frame(voutFrame, background='#A4E1FF')
        voutOptFrame.grid(row=0, column=0)
        voutOptRadioButton1 = tk.Radiobutton(
            voutOptFrame,
            text='Fixed Steps',
            variable=self.varVoutOpt,
            value='fixed',
            command=self.validate_vout_opt,
            background='#A4E1FF',
        )
        voutOptRadioButton2 = tk.Radiobutton(
            voutOptFrame,
            text='Increment Steps',
            variable=self.varVoutOpt,
            value='incr',
            command=self.validate_vout_opt,
            background='#A4E1FF'
        )
        voutOptRadioButton1.grid(row=1, column=0, pady=2, sticky='w')
        voutOptRadioButton2.grid(row=2, column=0, pady=2, sticky='w')
        self.validate_vout_opt()
        #endregion

        #endregion

        #region Scope Channels
        scopeFrame = tk.Frame(self.setupLblFrame, background='#A4E1FF')
        scopeFrame.grid(row=0, column=6, rowspan=2, padx=2, pady=5, sticky='nw')
        self.varPWM = StringVar()
        lbl1 = tk.Label(scopeFrame, text='PWM Ch:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            scopeFrame,
            textvariable=self.varPWM,
            validate='focusout',
            validatecommand=self.validate_scope,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varPWM.set(ConfigMgr.testConditions101['scopePwmCh'])
        self.varSW = StringVar()
        lbl2 = tk.Label(scopeFrame, text='SW Ch:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            scopeFrame,
            textvariable=self.varSW,
            validate='focusout',
            validatecommand=self.validate_scope,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varSW.set(ConfigMgr.testConditions101['scopeSwCh'])
        self.varGateL = StringVar()
        lbl3 = tk.Label(scopeFrame, text='GateL Ch:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            scopeFrame,
            textvariable=self.varGateL,
            validate='focusout',
            validatecommand=self.validate_scope,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varGateL.set(ConfigMgr.testConditions101['scopeGateLCh'])
        self.varVdsh = StringVar()
        lbl4 = tk.Label(scopeFrame, text='VDSH Ch:', background='#A4E1FF')
        lbl4.grid(row=3, column=0, pady=2)
        entry4 = ttk.Entry(
            scopeFrame,
            textvariable=self.varVdsh,
            validate='focusout',
            validatecommand=self.validate_scope,
            width=4,
        )
        entry4.grid(row=3, column=1, pady=2)
        self.varVdsh.set(ConfigMgr.testConditions101['scopeVdshCh'])
        #endregion

        #region Soak Time
        soakFrame = tk.Frame(self.setupLblFrame, background='#A4E1FF')
        soakFrame.grid(row=1, column=8, padx=2, pady=0, sticky='nw')
        self.varSoakTime = StringVar()
        lbl1 = tk.Label(soakFrame, text='Soak Time:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            soakFrame,
            textvariable=self.varSoakTime,
            validate='focusout',
            validatecommand=self.validate_soak,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varSoakTime.set(ConfigMgr.testConditions101['soak'])
        lbl1unt = tk.Label(soakFrame, text='s', background='#A4E1FF')
        lbl1unt.grid(row=0, column=2, pady=2)
        #endregion

        #region Fall Image
        self.fallPersisFrame = tk.Frame(self.setupLblFrame, background='#A4E1FF')
        self.fallPersisFrame.grid(row=0, column=8, pady=2, sticky='nw')
        self.varFallImg = StringVar()
        self.varFallImg.set(ConfigMgr.testConditions101['fallImg'])
        fallImgChkBox = tk.Checkbutton(
            self.fallPersisFrame,
            text='Fall Img',
            variable=self.varFallImg,
            command=self.validate_fall_img,
            background='#A4E1FF',
        )
        fallImgChkBox.grid(row=0, column=0, pady=2, sticky='nw')
        #endregion

        #region Persistence time
        self.varPersis = StringVar()
        self.varPersis.set(ConfigMgr.testConditions101['persis'])
        persisLbl = tk.Label(self.fallPersisFrame, text='Persistence', background='#A4E1FF')
        persisLbl.grid(row=1, column=0, pady=2, sticky='nw')
        persistenceFrame = tk.Frame(self.fallPersisFrame, background='#A4E1FF')
        persistenceFrame.grid(row=2, column=0, pady=2, sticky='ne')
        persisEntry = tk.Entry(
            persistenceFrame,
            textvariable=self.varPersis,
            validate='focusout',
            validatecommand=self.validate_persistence,
            width=4,
        )
        persisEntry.grid(row=0, column=0, pady=2)
        persisUnt = tk.Label(persistenceFrame, text='s', background='#A4E1FF')
        persisUnt.grid(row=0, column=1, pady=2)
        #endregion

        # region test 1.01 Keithley Channels
        self.keithChsFrame = tk.LabelFrame(self.setupLblFrame, text='Keithley Channels', background='#A4E1FF')
        self.keithChsFrame.grid(row=3, column=0, columnspan=6, padx=2, pady=(0, 2), sticky='nw')
        self.kiin = StringVar()
        self.kiin.set(ConfigMgr.testConditions101['kiin'])
        self.kimon = StringVar()
        self.kimon.set(ConfigMgr.testConditions101['kimon'])
        self.kvout = StringVar()
        self.kvout.set(ConfigMgr.testConditions101['kvout'])
        self.ktmon = StringVar()
        self.ktmon.set(ConfigMgr.testConditions101['ktmon'])
        self.kvin = StringVar()
        self.kvin.set(ConfigMgr.testConditions101['kvin'])
        self.kiout = StringVar()
        self.kiout.set(ConfigMgr.testConditions101['kiout'])
        self.kioutr = StringVar()
        self.kioutr.set(ConfigMgr.testConditions101['kioutr'])
        self.kiinr = StringVar()
        self.kiinr.set(ConfigMgr.testConditions101['kiinr'])
        self.kvcc = StringVar()
        self.kvcc.set(ConfigMgr.testConditions101['kvcc'])
        self.kicc = StringVar()
        self.kicc.set(ConfigMgr.testConditions101['kicc'])
        self.kiccr = StringVar()
        self.kiccr.set(ConfigMgr.testConditions101['kiccr'])
        self.kvoutsw = StringVar()
        self.kvoutsw.set(ConfigMgr.testConditions101['kvoutsw'])
        self.kven = StringVar()
        self.kven.set(ConfigMgr.testConditions101['kven'])
        self.kien = StringVar()
        self.kien.set(ConfigMgr.testConditions101['kien'])
        self.kienr = StringVar()
        self.kienr.set(ConfigMgr.testConditions101['kienr'])
        self.kChBoxArgList = [
            self.keithChsFrame,
            ConfigMgr.testConditions101,
            validate.ps_keithley_channels,
            self.kvin,
            self.kiin,
            self.kiinr,
            self.kvout,
            self.kiout,
            self.kioutr,
            self.kvcc,
            self.kicc,
            self.kiccr,
            self.ktmon,
            self.kimon,
            self.kvoutsw,
            self.kven,
            self.kien,
            self.kienr,
        ]
        make.keithley_channel(self.kChBoxArgList)
        # endregion

        #region Temperature vs time soak between fsw sweeps
        self.timeFrame = tk.LabelFrame(self.setupLblFrame, text='CoolDown', background='#A4E1FF')
        self.timeFrame.grid(row=3, column=6, padx=2, pady=(1, 2), sticky='nw')
        self.varCoolDownTime = StringVar()
        timeRadioBtn = tk.Radiobutton(
            self.timeFrame,
            variable=self.varCoolDownTime,
            text='5 min',
            value='5min',
            command=self.validate_cooldown_time_opt,
            background='#A4E1FF',
        )
        timeRadioBtn.grid(row=0, column=0)
        timeRadioBtn2 = tk.Radiobutton(
            self.timeFrame,
            variable=self.varCoolDownTime,
            text='Temp.',
            value='temp',
            command=self.validate_cooldown_time_opt,
            background='#A4E1FF',
        )
        timeRadioBtn2.grid(row=1, column=0)
        self.varCoolDownTime.set(ConfigMgr.testConditions101['coolDOpt'])

        #endregion

        #region Inductor & Background selector
        # INDUCTOR
        self.inductorFrame = tk.Frame(self.setupLblFrame, background='#A4E1FF')
        self.inductorFrame.grid(row=3, column=7, rowspan=2, columnspan=2, padx=2, pady=(1, 2), sticky='nw')
        self.varInductorVal = StringVar()
        inducLbl = tk.Label(self.inductorFrame, text='Inductor', background='#A4E1FF')
        inducLbl.grid(row=0, column=0, pady=2)
        inducEntry = tk.Entry(
            self.inductorFrame,
            textvariable=self.varInductorVal,
            validate='focusout',
            validatecommand=lambda: validate.inductor([ConfigMgr.testConditions101, self.varInductorVal]),
            width=4,
        )
        inducEntry.grid(row=0, column=1, pady=2)
        inducUntLbl = tk.Label(self.inductorFrame, text='nH', background='#A4E1FF')
        inducUntLbl.grid(row=0, column=2, pady=2)

        # IMAGE BACKGROUND
        self.backgndFrame = tk.LabelFrame(self.inductorFrame, text='Img Backgnd', background='#A4E1FF')
        self.backgndFrame.grid(row=1, column=0, columnspan=3)
        self.varBackgndColor = StringVar()
        self.varBackgndColor.set('white')
        btn1 = tk.Radiobutton(
            self.backgndFrame,
            text='White',
            variable=self.varBackgndColor,
            value='white',
            command=lambda: validate.background_color([ConfigMgr.testConditions101, self.varBackgndColor]),
            background='#A4E1FF',
        )
        btn1.grid(row=0, column=0, pady=2)
        btn2 = tk.Radiobutton(
            self.backgndFrame,
            text='Black',
            variable=self.varBackgndColor,
            value='black',
            command=lambda: validate.background_color([ConfigMgr.testConditions101, self.varBackgndColor]),
            background='#A4E1FF',
        )
        btn2.grid(row=1, column=0, pady=2)
        #endregion

        #region Swithing Load Steps
        loadFrame2 = tk.LabelFrame(self.setupLblFrame, text='Switching Load Steps (A)', background='#A4E1FF')
        loadFrame2.grid(row=4, column=0, rowspan=3, columnspan=2, padx=2, pady=5, sticky='nw')
        self.varStartCurrent2 = StringVar()
        lbl1 = tk.Label(loadFrame2, text='Start Current:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            loadFrame2,
            textvariable=self.varStartCurrent2,
            validate='focusout',
            validatecommand=self.validate_current2,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartCurrent2.set(ConfigMgr.testConditions101['startCurr2'])
        self.varEndCurrent2 = StringVar()
        lbl2 = tk.Label(loadFrame2, text='End Current:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            loadFrame2,
            textvariable=self.varEndCurrent2,
            validate='focusout',
            validatecommand=self.validate_current2,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndCurrent2.set(ConfigMgr.testConditions101['endCurr2'])
        self.varStepCurrent2 = StringVar()
        lbl3 = tk.Label(loadFrame2, text='Step Current:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            loadFrame2,
            textvariable=self.varStepCurrent2,
            validate='focusout',
            validatecommand=self.validate_current2,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepCurrent2.set(ConfigMgr.testConditions101['stepCurr2'])
        #endregion

        #region RUN SELECTION
        runFrame = tk.LabelFrame(self.setupLblFrame, text='Run', background='#A4E1FF')
        runFrame.grid(row=4, column=2, rowspan=3, columnspan=2, padx=2, pady=5, sticky='nw')
        self.varRun1 = StringVar()
        self.varRun2 = StringVar()
        self.varRun3 = StringVar()
        self.varRun1.set(ConfigMgr.testConditions101['run1'])
        self.varRun2.set(ConfigMgr.testConditions101['run2'])
        self.varRun3.set(ConfigMgr.testConditions101['run3'])
        runArgList = [
            ConfigMgr.testConditions101,
            self.varRun1,
            self.varRun2,
            self.varRun3,
        ]
        chk1 = tk.Checkbutton(
            runFrame,
            text='Efficiency',
            variable=self.varRun1,
            command=lambda: validate.run(runArgList),
            background='#A4E1FF',
        )
        chk1.grid(row=0, column=0, pady=2)
        chk2 = tk.Checkbutton(
            runFrame,
            text='Switching',
            variable=self.varRun2,
            command=lambda: validate.run(runArgList),
            background='#A4E1FF',
        )
        chk2.grid(row=1, column=0, pady=2)
        chk3 = tk.Checkbutton(
            runFrame,
            text='Snapshot',
            variable=self.varRun3,
            command=lambda: validate.run(runArgList),
            background='#A4E1FF',
        )
        chk3.grid(row=2, column=0, pady=2)
        #endregion

        #region snapshot options
        snapOptFrame = tk.LabelFrame(self.setupLblFrame, text='Snapshot Options', background='#A4E1FF')
        snapOptFrame.grid(row=4, column=3, rowspan=3, columnspan=2, padx=2, pady=5, sticky='nw')
        self.snapZoomShot = StringVar()
        self.snapZoomShot.set(ConfigMgr.testConditions101['snapScopeZoom'])
        self.loadTimeOff = StringVar()
        self.loadTimeOff.set(ConfigMgr.testConditions101['loadRest'])
        self.loadRest = StringVar()
        self.loadRest.set(ConfigMgr.testConditions101['loadRestTime'])
        self.snapPause = StringVar()
        self.snapPause.set(ConfigMgr.testConditions101['snapPause'])
        snapOptArgs = [
            ConfigMgr.testConditions101,
            self.snapZoomShot,
            self.loadTimeOff,
            self.loadRest,
            self.snapPause,
        ]
        chk1 = tk.Checkbutton(
            snapOptFrame,
            text='second shot',
            variable=self.snapZoomShot,
            command=lambda: validate.snap_opt(snapOptArgs),
            background='#A4E1FF',
        )
        chk1.grid(row=0, column=0, pady=2, sticky='w')
        tip1 = CreateToolTip(chk1, '''Enables a second screenshot. The GUI will pause and wait for user to adjust zoom or cursor or any changes to save the next screenshot''')
        chk2 = tk.Checkbutton(
            snapOptFrame,
            text='load off time',
            variable=self.loadTimeOff,
            command=lambda: validate.snap_opt(snapOptArgs),
            background='#A4E1FF',
        )
        chk2.grid(row=1, column=0, pady=2, sticky='w')
        tip2 = CreateToolTip(chk2, '''Turns off the load for this amount of time between load steps''')
        loadRestEntry = tk.Entry(
            snapOptFrame,
            textvariable=self.loadRest,
            validate='focusout',
            validatecommand=lambda: validate.snap_opt(snapOptArgs),
            width=4,
        )
        loadRestEntry.grid(row=1, column=1, pady=2, sticky='w')
        loadRestUnt = tk.Label(snapOptFrame, text='s', background='#A4E1FF')
        loadRestUnt.grid(row=1, column=2, pady=2, sticky='w')
        chk3 = tk.Checkbutton(
            snapOptFrame,
            text='pause',
            variable=self.snapPause,
            command=lambda: validate.snap_opt(snapOptArgs),
            background='#A4E1FF',
        )
        chk3.grid(row=2, column=0, pady=2, sticky='w')
        tip3 = CreateToolTip(chk3, '''Select between soak time and pause. When selected it will pause instead of waiting for soak time''')
        #endregion

        #region SW Only Fall dead time
        self.swOnlyFrame = tk.LabelFrame(self.setupLblFrame, text='Fall DT Config', background='#A4E1FF')
        self.swOnlyFrame.grid(row=4, column=3, rowspan=3, columnspan=2, padx=2, pady=5, sticky='nw')
        self.varRefLevel = StringVar()
        self.varCrossNum = StringVar()
        self.varRefLevel.set(ConfigMgr.testConditions101['refLevel'])
        self.varCrossNum.set(ConfigMgr.testConditions101['crossNum'])
        swOnlyArgList = [
            self.swOnlyFrame,
            ConfigMgr.testConditions101,
            self.varRefLevel,
            self.varCrossNum,
        ]
        # REF LEVEL
        refLevelLbl = tk.Label(self.swOnlyFrame, text='Crossing Voltage Level:', background='#A4E1FF')
        refLevelLbl.grid(row=0, column=0, padx=2, pady=2, sticky='e')
        refLevelEntry = tk.Entry(
            self.swOnlyFrame,
            textvariable=self.varRefLevel,
            validate='focusout',
            validatecommand=lambda: validate.sw_only_fall_dt(swOnlyArgList),
            width=6,
        )
        refLevelEntry.grid(row=0, column=1, padx=2, pady=2)
        refLevelUnt = tk.Label(self.swOnlyFrame, text='V', background='#A4E1FF')
        refLevelUnt.grid(row=0, column=2, padx=2, pady=2)
        # NUMBER OF CROSSINGS
        crossNumLbl = tk.Label(self.swOnlyFrame, text='Number of Crossing:', background='#A4E1FF')
        crossNumLbl.grid(row=1, column=0, padx=2, pady=2, sticky='e')
        crossNumEntry = tk.Entry(
            self.swOnlyFrame,
            textvariable=self.varCrossNum,
            validate='focusout',
            validatecommand=lambda: validate.sw_only_fall_dt(swOnlyArgList),
            width=6,
        )
        crossNumEntry.grid(row=1, column=1, padx=2, pady=2)
        self.validate_scope()
        #endregion

    def graph_box(self):
        graphF = ttk.Frame(self.graphFrame)
        graphF.grid(row=0, column=0)

        plotCanvas = FigureCanvasTkAgg(self.f, master=graphF)
        plotCanvas.draw()
        plotCanvas.get_tk_widget().grid(row=0, column=0, padx=2, pady=2)

    def table_box(self):
        self.table['columns'] = (
            'Date',
            'Time',
            'Vin',
            'Iin',
            'Vcc',
            'Icc',
            'Vout',
            'Iout',
            'Efficiency',
            'Temp',
            'Humidity',
        )
        self.table.column('#0', width=0, stretch=tk.NO)
        self.table.column('Date', anchor=tk.W, width=25)
        self.table.column('Time', anchor=tk.W, width=25)
        self.table.column('Vin', anchor=tk.CENTER, width=25)
        self.table.column('Iin', anchor=tk.CENTER, width=25)
        self.table.column('Vcc', anchor=tk.CENTER, width=25)
        self.table.column('Icc', anchor=tk.CENTER, width=25)
        self.table.column('Vout', anchor=tk.CENTER, width=25)
        self.table.column('Iout', anchor=tk.CENTER, width=25)
        self.table.column('Efficiency', anchor=tk.CENTER, width=30)
        self.table.column('Temp', anchor=tk.CENTER, width=25)
        self.table.column('Humidity', anchor=tk.CENTER, width=30)

        self.table.heading('#0', anchor=tk.W, text='')
        self.table.heading('Date', anchor=tk.CENTER, text='Date')
        self.table.heading('Time', anchor=tk.CENTER, text='Time')
        self.table.heading('Vin', anchor=tk.CENTER, text='Vin')
        self.table.heading('Iin', anchor=tk.CENTER, text='Iin')
        self.table.heading('Vcc', anchor=tk.CENTER, text='Vcc')
        self.table.heading('Icc', anchor=tk.CENTER, text='Icc')
        self.table.heading('Vout', anchor=tk.CENTER, text='Vout')
        self.table.heading('Iout', anchor=tk.CENTER, text='Iout')
        self.table.heading('Efficiency', anchor=tk.CENTER, text='Efficiency')
        self.table.heading('Temp', anchor=tk.CENTER, text='Temp.')
        self.table.heading('Humidity', anchor=tk.CENTER, text='Humidity')

        self.table.tag_configure('oddrow', background='white')
        self.table.tag_configure('evenrow', background='lightblue')

    #region VALIDATION FUNCTIONS
    def validate_current(self):
        ConfigMgr.testConditions101['startCurr'] = self.varStartCurrent.get()
        ConfigMgr.testConditions101['endCurr'] = self.varEndCurrent.get()
        ConfigMgr.testConditions101['stepCurr'] = self.varStepCurrent.get()
        return True

    def validate_current2(self):
        ConfigMgr.testConditions101['startCurr2'] = self.varStartCurrent2.get()
        ConfigMgr.testConditions101['endCurr2'] = self.varEndCurrent2.get()
        ConfigMgr.testConditions101['stepCurr2'] = self.varStepCurrent2.get()
        return True

    def validate_fsw(self):
        ConfigMgr.testConditions101['startFsw'] = self.varStartFsw.get()
        ConfigMgr.testConditions101['endFsw'] = self.varEndFsw.get()
        ConfigMgr.testConditions101['stepFsw'] = self.varStepFsw.get()
        ConfigMgr.testConditions101['fsw1'] = self.varFsw01.get()
        ConfigMgr.testConditions101['fsw2'] = self.varFsw02.get()
        ConfigMgr.testConditions101['fsw3'] = self.varFsw03.get()
        ConfigMgr.testConditions101['fsw4'] = self.varFsw04.get()
        ConfigMgr.testConditions101['fsw5'] = self.varFsw05.get()
        ConfigMgr.testConditions101['fsw6'] = self.varFsw06.get()
        return True

    def validate_fsw_opt(self):
        ConfigMgr.testConditions101['fswOpt'] = self.varFswOpt.get()
        if self.varFswOpt.get() == 'incr':
            self.incrFreqFrame.grid()
            self.fswFixedFrame.grid_remove()
        else:
            self.incrFreqFrame.grid_remove()
            self.fswFixedFrame.grid()
        return True

    def validate_vin(self):
        ConfigMgr.testConditions101['startPVin'] = self.varStartVin.get()
        ConfigMgr.testConditions101['endPVin'] = self.varEndVin.get()
        ConfigMgr.testConditions101['stepPVin'] = self.varStepVin.get()
        ConfigMgr.testConditions101['PVin1'] = self.varVin01.get()
        ConfigMgr.testConditions101['PVin2'] = self.varVin02.get()
        ConfigMgr.testConditions101['PVin3'] = self.varVin03.get()
        ConfigMgr.testConditions101['PVin4'] = self.varVin04.get()
        ConfigMgr.testConditions101['PVin5'] = self.varVin05.get()
        ConfigMgr.testConditions101['PVin6'] = self.varVin06.get()
        return True

    def validate_vin_opt(self):
        ConfigMgr.testConditions101['pvinOpt'] = self.varVinOpt.get()
        if self.varVinOpt.get() == 'incr':
            self.incrFrame.grid()
            self.fixedFrame.grid_remove()
        else:
            self.incrFrame.grid_remove()
            self.fixedFrame.grid()
        return True

    def validate_vout(self):
        ConfigMgr.testConditions101['startVout'] = self.varStartVout.get()
        ConfigMgr.testConditions101['endVout'] = self.varEndVout.get()
        ConfigMgr.testConditions101['stepVout'] = self.varStepVout.get()
        ConfigMgr.testConditions101['vout1'] = self.varVout01.get()
        ConfigMgr.testConditions101['vout2'] = self.varVout02.get()
        ConfigMgr.testConditions101['vout3'] = self.varVout03.get()
        ConfigMgr.testConditions101['vout4'] = self.varVout04.get()
        ConfigMgr.testConditions101['vout5'] = self.varVout05.get()
        ConfigMgr.testConditions101['vout6'] = self.varVout06.get()
        return True

    def validate_vout_opt(self):
        ConfigMgr.testConditions101['voutOpt'] = self.varVoutOpt.get()
        if self.varVoutOpt.get() == 'incr':
            self.incrVoutFrame.grid()
            self.fixedVoutFrame.grid_remove()
        else:
            self.incrVoutFrame.grid_remove()
            self.fixedVoutFrame.grid()
        return True

    def validate_vcc(self):
        ConfigMgr.testConditions101['startVcc'] = self.varStartVcc.get()
        ConfigMgr.testConditions101['endVcc'] = self.varEndVcc.get()
        ConfigMgr.testConditions101['stepVcc'] = self.varStepVcc.get()
        return True

    def validate_soak(self):
        ConfigMgr.testConditions101['soak'] = self.varSoakTime.get()
        return True

    def validate_scope(self):
        ConfigMgr.testConditions101['scopePwmCh'] = self.varPWM.get()
        ConfigMgr.testConditions101['scopeSwCh'] = self.varSW.get()
        ConfigMgr.testConditions101['scopeGateLCh'] = self.varGateL.get()
        ConfigMgr.testConditions101['scopeVdshCh'] = self.varVdsh.get()
        if self.varGateL.get() == '':
            self.swOnlyFrame.grid()
        else:
            self.swOnlyFrame.grid_remove()
        return True

    def validate_fall_img(self):
        ConfigMgr.testConditions101['fallImg'] = self.varFallImg.get()
        return True

    def validate_persistence(self):
        ConfigMgr.testConditions101['persis'] = self.varPersis.get()
        return True

    def validate_cooldown_time_opt(self):
        ConfigMgr.testConditions101['coolDOpt'] = self.varCoolDownTime.get()
        return True
    #endregion

    #region RUN TEST
    def run_tests(self):
        self.running = KillableThread(name='test', target=self._run_tests)
        self.running.start()

    def _run_tests(self):
        self.buttons_switch()
        self.run_efficiency()
        self.buttons_switch()
        return

    def kill_test(self):
        self.running.raise_exception()
        self.running.join()
        self.buttons_switch()

    def buttons_switch(self):
        if(str(self.runButton.cget('state')) == 'normal'):
            self.runButton['state'] = 'disabled'
            self.killBtn['state'] = 'normal'
        else:
            self.runButton['state'] = 'normal'
            self.killBtn['state'] = 'disabled'

    def animate_graph(self, interval):
        print('animate_graph called!')
        self.a.clear()
        self.a.plot(self.xAxis, self.yAxis)

    def run_efficiency(self):
        try:
            self.tempList = list()
            self.currList = list()
            self.currList2 = list()
            self.freqList = list()
            self.pvinList = list()
            self.voutList = list()
            self.vccList = list()
            self.modeList = list()
            self.enList = list()
            self.biasList = list()
            self.delayList = list()
            self.statusLabel['text'] = 'Initializing test Variables...'
            BACKGROUND = ConfigMgr.testConditions101['background']
            aux.setting_pre_test_variables(
                ConfigMgr.testConditions101,
                ConfigMgr.tempSteps,
                self.tempList,
                self.currList,
                self.currList2,
                self.freqList,
                self.pvinList,
                self.voutList,
                self.vccList,
                self.modeList,
                self.enList,
                self.biasList,
                self.delayList
            )
            self.statusLabel['text'] = 'Creating folder to save data'
            folderName = aux.make_dir('efficiency')
            self.statusLabel['text'] = 'Initializing Instruments...'
            if self.load is None:
                self.statusLabel['text'] = 'Initializing Load...'
                self.load = init.eload()
            if self.keith is None:
                self.statusLabel['text'] = 'Initializing Keithley...'
                self.keith = init.keith()
            if self.vcc is None:
                self.statusLabel['text'] = 'Initializing VCC...'
                self.vcc = init.vcc()
            if self.en is None:
                self.statusLabel['text'] = 'Initializing EN...'
                self.en = init.enable()
            if self.vin is None:
                self.statusLabel['text'] = 'Initializing VIN...'
                self.vin = init.vin()
            if self.scope is None:
                self.statusLabel['text'] = 'Initializing OSCILLOSCOPE...'
                if eval(ConfigMgr.testConditions101['run1']) or eval(ConfigMgr.testConditions101['run2']):
                    self.scope = init.oscilloscope()
                else:
                    self.scope = self.init_scope()

            test101Dict = {}
            inductorVal = ConfigMgr.testConditions101['inductorVal']
            self.statusLabel['text'] = 'Creating Excel file...'
            wb = self.create_excel(title='Efficiency', index=0, config=ConfigMgr.testConditions101)
            ws1 = wb['Efficiency']
            self.data101_excel(ws1)
            for item in self.table.get_children():
                self.table.delete(item)
            self.xAxis = list()
            self.yAxis = list()
            self.statusLabel['text'] = 'Clearing Plots...'
            self.a.clear()
            self.a.plot(self.xAxis, self.yAxis)
            self.plotCanvas.draw_idle()

            # LOAD Set-up
            if self.load is not None:
                self.statusLabel['text'] = 'Setting load to CCH...'
                self.load.set_channel(index=ConfigMgr.instr['loadChannel'][-1])
                self.load.set_mode(mode='CCH')
                self.load.set_current('STATIC', 0)
                self.load.output('OFF')

            if eval(ConfigMgr.testConditions101['run1']):
                # Keith Set-up
                if self.keith is not None:
                    self.statusLabel['text'] = 'Setting Keithley channels...'
                    self.keith.reset_keithley()
                    self.keith.select_function(function='VOLT:DC', channel='(@101:120)')
                    self.keith.select_range(function='VOLT:DC', range='AUTO ON', channel='(@101:120)')
                    self.keith.set_apperture(function='VOLT:DC', value='20e-3', channel='(@101:120)')
                    self.keith.route_to_scan('(@101:120)')
                    self.keith.enable_scan('ON')
                    self.keith.number_of_channels_to_scan(20)
                    self.keith.data_elements(reading=1)

            # VCC Set-up
            if eval(ConfigMgr.instr['vccPSOnOff']):
                self.statusLabel['text'] = 'Setting VCC start level...'
                if self.vcc is not None:
                    self.vcc.vdrvCh.Configure_VoltageLevel(Level=self.vccList[0], CurrentLimit=1)
                    self.vcc.vdrvCh.Enable(True)

            # EN Set-up
            if eval(ConfigMgr.instr['enPSOnOff']):
                self.statusLabel['text'] = 'Setting VCC start level...'
                if self.en is not None:
                    self.en.ch.Configure_VoltageLevel(Level=self.enList[0], CurrentLimit=1)
                    self.en.ch.Enable(True)

            sleep(0.5)

            # VIN Set-up
            if self.vin is not None:
                self.statusLabel['text'] = 'Setting VIN start level...'
                self.vin.set_over_voltage_protection(voltage=20)
                self.vin.ch.Configure_VoltageLevel(Level=self.pvinList[0], CurrentLimit=30)
                self.vin.ch.Enable(True)

            sleep(0.5)

            if self.dongle is None:
                self.statusLabel['text'] = 'Initializing Dongle...'
                self.dongle = init.dongles()

            # SCOPE Set-up
            if eval(ConfigMgr.testConditions101['run1']) or eval(ConfigMgr.testConditions101['run2']):
                if self.scope is not None:
                    self.statusLabel['text'] = 'Setting OSCILLOSCOPE initial config...'
                    try:
                        ch1 = self.scope.GetChannel(Index=1)
                        ch1.Enable(False)
                    except Exception as e:
                        print('Failed Creating Scope channel Object')
                        pass
                    try:
                        swCh = eval(ConfigMgr.testConditions101['scopeSwCh'])
                        self.scopeSwCh = self.scope.GetChannel(Index=swCh)
                        self.statusLabel['text'] = 'Turning SW Channel on...'
                        self.scopeSwCh.Enable(True)
                        self.scope.set_attenuation(swCh, eval(ConfigMgr.instr[f'scopeAttnCh{swCh}']))
                    except:
                        if ConfigMgr.testConditions101['scopeSwCh'] == '':
                            print('SW Channel off')
                        else:
                            print('Failed Creating SW Scope channel Object')
                        swCh = 0
                        self.scopeSwCh = None
                        pass
                    try:
                        pwmCh = eval(ConfigMgr.testConditions101['scopePwmCh'])
                        self.scopePwmCh = self.scope.GetChannel(Index=pwmCh)
                        self.statusLabel['text'] = 'Turning PWM Channel on...'
                        self.scopePwmCh.Enable(True)
                        self.scope.set_attenuation(pwmCh, eval(ConfigMgr.instr[f'scopeAttnCh{pwmCh}']))
                    except:
                        if ConfigMgr.testConditions101['scopePwmCh'] == '':
                            print('PWM Channel off')
                        else:
                            print('Failed Creating PWM Scope channel Object')
                        pwmCh = 0
                        self.scopePwmCh = None
                        pass
                    try:
                        gateLCh = eval(ConfigMgr.testConditions101['scopeGateLCh'])
                        self.scopeGateLCh = self.scope.GetChannel(Index=gateLCh)
                        self.statusLabel['text'] = 'Turning GATEL Channel on...'
                        self.scopeGateLCh.Enable(True)
                        self.scope.set_attenuation(gateLCh, eval(ConfigMgr.instr[f'scopeAttnCh{gateLCh}']))
                    except:
                        if ConfigMgr.testConditions101['scopeGateLCh'] == '':
                            print('Gate L Channel off')
                        else:
                            print('Failed Creating GateL Scope channel Object')
                        gateLCh = 0
                        self.scopeGateLCh = None
                        pass
                    try:
                        vdshCh = eval(ConfigMgr.testConditions101['scopeVdshCh'])
                        self.scopeVdshCh = self.scope.GetChannel(Index=vdshCh)
                        self.statusLabel['text'] = 'Turning VDSH Channel on...'
                        self.scopeVdshCh.Enable(True)
                        self.scope.set_attenuation(vdshCh, eval(ConfigMgr.instr[f'scopeAttnCh{vdshCh}']))
                    except:
                        if ConfigMgr.testConditions101['scopeVdshCh'] == '':
                            print('VDSH Channel off')
                        else:
                            print('Failed Creating VDSH Scope channel Object')
                        self.scopeVdshCh = None
                        vdshCh = 0
                        pass

                    # Overlay display
                    self.scope.display_overlay()
                    self.scope.horizontal_delay('OFF')
                    self.scope.set_persistence('1')
                    self.scope.set_acquisition_mode('HIR')

                    # Config Probe Channels
                    if self.scopePwmCh is not None:
                        self.statusLabel['text'] = 'PWM channel position, scale, attenuation, impedance, coupling'
                        self.scopePwmCh.ProbeSetup(
                            Coupling="DC",
                            Bandwidth=None,
                            Vrange=int(eval(ConfigMgr.instr[f'scopeVertScaleCh{pwmCh}']) * 10),
                            Offset=0,
                            Position=0,
                        )
                        self.scope.set_attenuation(ch=pwmCh, value=eval(ConfigMgr.instr[f'scopeAttnCh{pwmCh}']))
                        self.scope.set_termination(ch=pwmCh, value=eval(ConfigMgr.instr[f'scopeTermCh{pwmCh}']))
                    if self.scopeSwCh is not None:
                        self.statusLabel['text'] = 'SW channel position, scale, attenuation, impedance, coupling'
                        self.scopeSwCh.ProbeSetup(
                            Coupling="DC",
                            Bandwidth=None,
                            Vrange=int(eval(ConfigMgr.instr[f'scopeVertScaleCh{swCh}']) * 10),
                            Offset=0,
                            Position=0,
                        )
                        self.scope.set_termination(ch=swCh, value=eval(ConfigMgr.instr[f'scopeTermCh{swCh}']))
                        self.scope.set_attenuation(ch=swCh, value=eval(ConfigMgr.instr[f'scopeAttnCh{swCh}']))
                    if self.scopeGateLCh is not None:
                        self.statusLabel['text'] = 'GATEL channel position, scale, attenuation, impedance, coupling'
                        self.scopeGateLCh.ProbeSetup(
                            Coupling="DC",
                            Bandwidth=None,
                            Vrange=int(eval(ConfigMgr.instr[f'scopeVertScaleCh{gateLCh}']) * 10),
                            Offset=0,
                            Position=0,
                        )
                        self.scope.set_attenuation(ch=gateLCh, value=eval(ConfigMgr.instr[f'scopeAttnCh{gateLCh}']))
                        self.scope.set_termination(ch=gateLCh, value=eval(ConfigMgr.instr[f'scopeTermCh{gateLCh}']))
                    if self.scopeVdshCh is not None:
                        self.statusLabel['text'] = 'VDSH channel position, scale, attenuation, impedance, coupling'
                        self.scopeVdshCh.ProbeSetup(
                            Coupling="DC",
                            Bandwidth=None,
                            Vrange=int(eval(ConfigMgr.instr[f'scopeVertScaleCh{vdshCh}']) * 10),
                            Offset=0,
                            Position=0,
                        )
                        self.scope.set_attenuation(ch=vdshCh, value=eval(ConfigMgr.instr[f'scopeAttnCh{vdshCh}']))
                        self.scope.set_termination(ch=vdshCh, value=eval(ConfigMgr.instr[f'scopeTermCh{vdshCh}']))

                    # Trigger
                    self.statusLabel['text'] = 'Setting Trigger'
                    self.scope.Trigger_Edge(
                        Level=2.00,
                        Slope='RISE',
                        Position=2,
                        Coupling='DC',
                        ChannelIndex=int(swCh)
                    )
                    self.scope.trigger_mode('AUTO')

                    # Vertical Position
                    self.scope.verpos(swCh, str(ConfigMgr.instr['scopeVertPosCh' + str(swCh)]))
                    self.scope.set_channel_label(swCh, 'SW', '50', '-30')

                    if gateLCh:
                        self.scope.verpos(gateLCh, str(ConfigMgr.instr['scopeVertPosCh' + str(gateLCh)]))
                        self.scope.set_channel_label(gateLCh, 'GateL', '50', '-30')

                    if pwmCh:
                        self.scope.verpos(pwmCh, str(ConfigMgr.instr['scopeVertPosCh' + str(pwmCh)]))
                        self.scope.set_channel_label(pwmCh, 'PWM', '50', '-30')

                    if vdshCh:
                        self.scope.verpos(vdshCh, str(ConfigMgr.instr['scopeVertPosCh' + str(vdshCh)]))
                        self.scope.set_channel_label(vdshCh, 'VDSH', '50', '-30')

                    # Horizontal Scale
                    self.statusLabel['text'] = 'Setting Horizontal Scale'
                    self.scope.horscale('4e-9')
                    self.scope.horizontal_mode('AUTO')
                    self.scope.horizontal_sample_rate('5e+9')
                    self.scope.horpos(50)

                    # Measurements
                    self.statusLabel['text'] = 'Adding frequency measurement for SW Channel'
                    self.scope.set_measurement(
                        MeasureIndex=5,
                        MeasureType='FREQUENCY',
                        State='ON',
                        Source=swCh,
                    )

                    # Scope Run
                    self.scope.Arm(Continuous=True)
                    try:
                        self.scope.set_persistence(ConfigMgr.testConditions101['persis'])
                    except:
                        self.scope.set_persistence('OFF')
                    self.scope.vertical_cursor_on()
                    self.scope.set_cursor_function('WAVEFORM')
                    self.scope.cursor_split()
                    self.scope.waveform_cursor_to_channel(cursor=1, ch=swCh)
                    self.scope.waveform_cursor_to_channel(cursor=2, ch=swCh)

            # Set the first row for data output...
            RowIndex = 3
            initRowIndex2 = 3
            RowIndex2 = 3
            tableID = 0
            sw_freq = 0
            deadRise = 0
            deadFall = 0
            sw_max = 0
            sw_rise = 0
            sw_fall = 0
            gatel_rise = 0
            gatel_fall = 0
            loadOn = 0
            bomtxt = f'Family: {ConfigMgr.instr["psBoardFamily"]}@&Model: {ConfigMgr.instr["psBoardModel"]}@&Revision: {ConfigMgr.instr["psBoardSilicRev"]}'
            with open(folderName + f'/data/efficiency.csv', '+a') as csvFile:
                csvFile.write(f'{bomtxt},,{len(self.currList)},23\n')
                csvFile.write(f'date,time,Column Title,Load,Measured Vin (V),Measured Iin (mV),Measured Iin (A),Measured VCC (V),Measured Icc (mV),Measured Icc (A),Measured EN (V),Measured Ien (mV),Measured Ien (A),Measured Vout (V),Measured Iout (mV),Measured Iout (A),TMON (V),Vout SW (V),Power Vin (W),Power Vcc (W),Power out (W),Power Loss (W),Efficiency (%),Temperature (C),Imon (V),Rise Dead Time (ns),Fall Dead Time (ns)\n')

            with open(folderName + f'/data/switch.csv', '+a') as csvFile:
                csvFile.write(f'{bomtxt},,{len(self.currList2)},8\n')
                csvFile.write(f'date,time,Column Title,Load,SW Max (V),SW Rise Time (ns),SW Fall Time (ns),Gate Low Rise Time (ns),Gate Low Fall Time (ns),VDSH Max (V),SW Ringing Freq. (MHz),Max GL CdVdt (V)\n')

            for t in self.tempList:
                if eval(ConfigMgr.instr['thermOnOff']):
                    self.statusLabel['text'] = f'Setting Temperature to {t}\u00b0C'
                    self.therm.set_temperature(t)
                    sleep(6)
                    currTemp = self.therm.get_temperature()
                    currTemp = float(currTemp)
                    while (abs(currTemp - t) > 1):
                        currTemp = self.therm.get_temperature()
                        currTemp = float(currTemp)
                        sleep(60)
                else:
                    messagebox.showwarning('Temperature', f'Set Temperature to {t}\u00B0C')

                for EN in self.enList:
                    if self.en is not None:
                        self.statusLabel['text'] = f'Setting Vcc to {EN}V'
                        self.en.ch.Configure_VoltageLevel(Level=EN, CurrentLimit=1)
                        self.en.ch.Enable(True)
                    else:
                        messagebox.showwarning('EN', f'Set Enable to {EN}V')

                    dataVcc = list()
                    for vcc in self.vccList:
                        if self.vcc is not None:
                            self.statusLabel['text'] = f'Setting Vcc to {vcc}V'
                            self.vcc.vdrvCh.Configure_VoltageLevel(Level=vcc, CurrentLimit=1)
                            self.vcc.vdrvCh.Enable(True)
                        else:
                            messagebox.showwarning('Vcc', f'Set Vcc to {vcc}V')

                        firstRun = 1
                        dataF = list()
                        for f in self.freqList:
                            # Changing Fsw
                            if self.dongle is not None:
                                self.statusLabel['text'] = f'Setting Frequency to {f}kHz'
                                self.dongle.set_frequency(freq=f)
                            else:
                                messagebox.showwarning('Frequency', f'Set Frequency to {f}kHz')

                            dataVout = list()
                            for vout in self.voutList:
                                # Changing Vout
                                if self.dongle is not None:
                                    self.statusLabel['text'] = f'Setting Vout to {vout}V'
                                    self.dongle.set_vout(voltage=float(vout))
                                else:
                                    messagebox.showwarning('Vout', f'Set Vout to {vout}V')

                                dataVin = list()
                                for vin in self.pvinList:
                                    # Changing Vin
                                    if eval(ConfigMgr.instr['vinPSOnOff']):
                                        if self.vin is not None:
                                            self.statusLabel['text'] = f'Setting Vin to {vin}V'
                                            self.vin.ch.Configure_VoltageLevel(Level=vin, CurrentLimit=30)
                                            self.vin.ch.Enable(True)
                                    else:
                                        messagebox.showwarning('Vin', f'Set Vin to {vin}V')

                                    dataList = list()
                                    deadRise = 0
                                    deadFall = 0
                                    if (not firstRun) and (ConfigMgr.testConditions101['coolDOpt'] == '5min'):
                                        cooldownTime = 300
                                        while cooldownTime > 0:
                                            sleep(1)
                                            cooldownTime -= 1
                                            ctmin = int(cooldownTime // 60)
                                            ctsec = int(cooldownTime % 60)
                                            self.statusLabel['text'] = f'Cooling down {ctmin:2d}:{ctsec:02d}'
                                    elif (not firstRun):
                                        temp = self.dongle.get_temp()
                                        while (temp > self.coolDownTemp):
                                            sleep(1)
                                            temp = self.dongle.get_temp()
                                            self.statusLabel['text'] = f'Waiting to cool down, temp. = {temp}\u00b0C'
                                    firstRun = 0
                                    if eval(ConfigMgr.testConditions101['run1']):
                                        for i in self.currList:
                                            mdate = dt.datetime.now()
                                            mydate = mdate.strftime('%m/%d/%Y')
                                            mytime = mdate.strftime('%H-%M-%S')
                                            if self.scope is not None:
                                                self.scope.run()  # Run
                                                self.scope.Trigger_Edge(
                                                    Level=2.00,
                                                    Slope='RISE',
                                                    Position=2,
                                                    Coupling='DC',
                                                    ChannelIndex=int(swCh)
                                                )
                                                self.scope.trigger_mode('AUTO')
                                                self.scope.horscale('4e-9')
                                                self.scope.set_vertical_scale(swCh, ConfigMgr.instr[f'scopeVertScaleCh{swCh}'])

                                            # Set the current
                                            if self.load is not None:
                                                if not loadOn:
                                                    self.load.set_channel(ConfigMgr.instr['loadChannel'][-1])
                                                    self.load.set_mode('CCH')
                                                self.statusLabel['text'] = f'Setting Load to {i}A'
                                                self.load.set_current('STATIC', i)
                                                if i == 0:
                                                    self.load.output('OFF')
                                                elif not loadOn:
                                                    self.load.output('ON')
                                                    loadOn = 1

                                            startTime = dt.datetime.now()
                                            if self.scope is not None:
                                                self.scope.vertical_cursor_on()
                                                self.scope.set_screen_text(text=f"{inductorVal}nH, {f}kHz, {vcc}Vcc, {vout}Vout, {vin}Vin, {i}A", text_no=1, xpos="1", ypos="8", FontSize=12)

                                                pointA = 0
                                                pointB = 0
                                                if gateLCh:
                                                    self.scope.horpos(50)
                                                    try:
                                                        self.scope.set_persistence(ConfigMgr.testConditions101['persis'])
                                                    except:
                                                        self.scope.set_persistence('OFF')
                                                    self.scope.reset_statistics()
                                                    try:
                                                        if eval(ConfigMgr.testConditions101['persis']) < 1:
                                                            sleep(1)
                                                        else:
                                                            sleep(eval(ConfigMgr.testConditions101['persis']))
                                                    except:
                                                        sleep(1)

                                                    # Stop Scope
                                                    self.scope.stop()
                                                    sleep(0.5)
                                                    gateLWave = self.scopeGateLCh.GetProbeWaveform(Timeout=0.5)
                                                    swWave = self.scopeSwCh.GetProbeWaveform(Timeout=0.5)
                                                    if self.scope.ID in ['4k', 'TekScope']:
                                                        gateLSlicedWave = gateLWave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                        pointA = gateLSlicedWave.Measurements_Utils.find_nth_crossing(1.0, "fall", 0)
                                                        swSlicedWave = swWave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                        pointB = swSlicedWave.Measurements_Utils.find_nth_crossing(1.0, "rise", 0)
                                                    else:
                                                        try:
                                                            pointA = gateLWave.Measurements_Utils.find_nth_crossing(1.0, 'fall', 0) - (gateLWave.index_to_time(-1) * 0.5)
                                                            pointB = swWave.Measurements_Utils.find_nth_crossing(1.0, 'rise', 0) - (swWave.index_to_time(-1) * 0.5)
                                                        except:
                                                            print('Unable to get dead time from oscilloscope. Possible reasons are: Probe is not connected or the part shut off')
                                                            pointA = 0
                                                            pointB = 0
                                                else:
                                                    self.scope.horpos(75)
                                                    self.scope.set_vertical_scale(swCh, 0.4)
                                                    self.scope.verpos(swCh, 2.5)
                                                    try:
                                                        self.scope.set_persistence(ConfigMgr.testConditions101['persis'])
                                                    except:
                                                        self.scope.set_persistence('OFF')
                                                    self.scope.reset_statistics()
                                                    try:
                                                        if eval(ConfigMgr.testConditions101['persis']) < 1:
                                                            sleep(1)
                                                        else:
                                                            sleep(eval(ConfigMgr.testConditions101['persis']))
                                                    except:
                                                        sleep(1)
                                                    # Stop Scope
                                                    self.scope.stop()
                                                    sleep(0.5)
                                                    wave = self.scopeSwCh.GetProbeWaveform(Timeout=0.5)
                                                    if self.scope.ID in ['4k', 'TekScope']:
                                                        slicedWave = wave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                        pointA = slicedWave.Measurements_Utils.argmin(return_index=True)
                                                        crossZeroA = aux.find_rise_dead_time(slicedWave, pointA, direction='BACKWARD')
                                                        crossZeroB = aux.find_rise_dead_time(slicedWave, pointA, direction='FORWARD')
                                                        if crossZeroA is None:
                                                            print('PointA for dead time rise not found')
                                                            crossZeroA = (0, 0, 0)
                                                        if crossZeroB is None:
                                                            print('PointB for dead time rise not found')
                                                            crossZeroB = (0, 0, 0)
                                                        pointA = crossZeroA[2]
                                                        pointB = crossZeroB[2]
                                                    else:
                                                        pointA = wave.Measurements_Utils.argmin(return_index=True)
                                                        crossZeroA = aux.find_rise_dead_time(wave, pointA, direction='BACKWARD')
                                                        crossZeroB = aux.find_rise_dead_time(wave, pointA, direction='FORWARD')
                                                        if crossZeroA is None:
                                                            print('PointA for dead time rise not found')
                                                            crossZeroA = (0, 0, 0)
                                                        if crossZeroB is None:
                                                            print('PointB for dead time rise not found')
                                                            crossZeroB = (0, 0, 0)
                                                        pointA = crossZeroA[2] - (wave.index_to_time(-1) * 0.75)
                                                        pointB = crossZeroB[2] - (wave.index_to_time(-1) * 0.75)
                                                    self.scope.run()

                                                if not pointA:
                                                    pointA = 0
                                                if not pointB:
                                                    pointB = 0

                                                self.scope.set_cursor_pos(cursor=1, pos=pointA)
                                                self.scope.set_cursor_pos(cursor=2, pos=pointB)
                                                deadRise = pointB - pointA
                                                sleep(0.5)

                                                # Waveform
                                                try:
                                                    fileName_i = folderName + f'/images/deadTime_rise_{t}C_{vcc}Vcc_{vout}Vout_{vin}Vin_{f}kHz_{i}A.png'
                                                    self.scope.saveimage(fileName_i, Background=BACKGROUND)
                                                except:
                                                    print(f'Error capturing deadTime_rise_{t}C_{vcc}Vcc_{vout}Vout_{vin}Vin_{f}kHz_{i}A image')

                                                self.scope.run()
                                                sleep(0.1)

                                                self.scope.Trigger_Edge(
                                                    Level=2.0,
                                                    Slope='FALL',
                                                    Position=2,
                                                    Coupling='DC',
                                                    ChannelIndex=int(swCh)
                                                )
                                                self.scope.trigger_mode('AUTO')

                                                if gateLCh:
                                                    self.scope.horpos(50)
                                                    try:
                                                        self.scope.set_persistence(ConfigMgr.testConditions101['persis'])
                                                    except:
                                                        self.scope.set_persistence('OFF')
                                                    self.scope.reset_statistics()
                                                    try:
                                                        if eval(ConfigMgr.testConditions101['persis']) < 1:
                                                            sleep(1)
                                                        else:
                                                            sleep(eval(ConfigMgr.testConditions101['persis']))
                                                    except:
                                                        sleep(1)

                                                    # Stop Scope
                                                    self.scope.stop()
                                                    sleep(0.5)
                                                    swWave = self.scopeSwCh.GetProbeWaveform(Timeout=0.5)
                                                    gateLWave = self.scopeGateLCh.GetProbeWaveform(Timeout=0.5)
                                                    if self.scope.ID == '4k':
                                                        swSlicedWave = swWave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                        pointA = swSlicedWave.Measurements_Utils.find_nth_crossing(1.0, "fall", 0)
                                                        gateLSlicedWave = gateLWave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                        pointB = gateLSlicedWave.Measurements_Utils.find_nth_crossing(2.0, "rise", 0)
                                                    else:
                                                        try:
                                                            pointA = swWave.Measurements_Utils.find_nth_crossing(1.0, 'fall', 0) - (swWave.index_to_time(-1) * 0.5)
                                                            pointB = gateLWave.Measurements_Utils.find_nth_crossing(2.0, 'rise', 0) - (gateLWave.index_to_time(-1) * 0.5)
                                                        except:
                                                            print('Unable to get dead time from oscilloscope. Possible reasons are: Probe is not connected or the part shut off')
                                                            pointA = 0
                                                            pointB = 0
                                                else:
                                                    try:
                                                        refLevel = float(ConfigMgr.testConditions101['refLevel'])
                                                    except:
                                                        refLevel = -0.2
                                                    try:
                                                        crossNum = int(ConfigMgr.testConditions101['crossNum']) - 1
                                                    except:
                                                        crossNum = 0
                                                    # Stop Scope
                                                    self.scope.horpos(25)
                                                    self.scope.set_vertical_scale(swCh, 0.4)
                                                    self.scope.verpos(swCh, 2.5)
                                                    try:
                                                        self.scope.set_persistence(ConfigMgr.testConditions101['persis'])
                                                    except:
                                                        self.scope.set_persistence('OFF')
                                                    self.scope.reset_statistics()
                                                    try:
                                                        if eval(ConfigMgr.testConditions101['persis']) < 1:
                                                            sleep(1)
                                                        else:
                                                            sleep(eval(ConfigMgr.testConditions101['persis']))
                                                    except:
                                                        sleep(1)
                                                    self.scope.stop()
                                                    sleep(0.5)
                                                    data = self.scopeSwCh.GetProbeWaveform(Timeout=0.5)
                                                    # data.save_to_file(folderName + f'/data/sw_fall_dt_{i}A.npz')
                                                    if self.scope.ID in ['4k', 'TekScope']:
                                                        slicedData = data.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                        firstZero = slicedData.Measurements_Utils.find_nth_crossing(0, 'fall', 0)
                                                        lowestPoint = slicedData.Measurements_Utils.argmin()
                                                        interestWave = slicedData.slice_by_time(start_time=lowestPoint, stop_time=20e-9)
                                                        A = np.vstack([interestWave.time, np.ones(len(interestWave.time))]).T
                                                        m, c = np.linalg.lstsq(A, interestWave.data, rcond=None)[0]
                                                        vertLine = 0
                                                        for val in (m * interestWave.time + c):
                                                            if val >= (refLevel):
                                                                idx = np.where((m * interestWave.time + c) == val)
                                                                vertLine = interestWave.time[idx]
                                                                break
                                                        # pointA = float(interestWave.time[0])
                                                        pointA = float(firstZero)
                                                        pointB = float(vertLine)
                                                    else:
                                                        if i < 18:
                                                            pointA = data.Measurements_Utils.argmin()
                                                            slicedWave = data.slice_by_time(start_time=pointA, stop_time=data.time[-1])
                                                            pointA = data.Measurements_Utils.find_nth_crossing(0, 'fall', 0)
                                                            pointB = slicedWave.Measurements_Utils.find_n_maxima(3)[0]
                                                        else:
                                                            pointA = data.Measurements_Utils.argmin()
                                                            slicedWave = data.slice_by_time(start_time=pointA, stop_time=data.time[-1])
                                                            pointA = data.Measurements_Utils.find_nth_crossing(0, 'fall', 0)
                                                            pointB = slicedWave.Measurements_Utils.find_n_maxima(3)[1]
                                                        try:
                                                            pointA = pointA - (data.index_to_time(-1) * 0.25)
                                                            pointB = pointB - (data.index_to_time(-1) * 0.25)
                                                        except:
                                                            print('Unable to get dead time from oscilloscope. Possible reasons are: Probe is not connected or the part shut off')
                                                            pointA = 0
                                                            pointB = 0
                                                    self.scope.run()

                                                if pointA is None:
                                                    pointA = 0
                                                if pointB is None:
                                                    pointB = 0

                                                self.scope.set_cursor_pos(cursor=1, pos=pointA)
                                                self.scope.set_cursor_pos(cursor=2, pos=pointB)
                                                sleep(0.5)
                                                deadFall = float(pointB - pointA)

                                                # Waveform
                                                try:
                                                    fileName_i = folderName + f'/images/deadTime_fall_{t}C_{vcc}Vcc_{vout}Vout_{vin}Vin_{f}kHz_{i}A.png'
                                                    if eval(ConfigMgr.testConditions101['fallImg']):
                                                        self.scope.saveimage(fileName_i, Background=BACKGROUND)
                                                except:
                                                    print(f'Error capturing deadTime_fall_{t}C_{vcc}Vcc_{vout}Vout_{vin}Vin_{f}kHz_{i}A image')

                                            endTime = dt.datetime.now()

                                            img_delay_time = endTime - startTime
                                            if img_delay_time.total_seconds() > (eval(ConfigMgr.testConditions101['soak']) - 1.75):
                                                pass
                                            else:
                                                sleep(eval(ConfigMgr.testConditions101['soak']) - img_delay_time.total_seconds() - 1.75)

                                            # keithley reading.
                                            if self.keith is not None:
                                                results = self.keith.read(num=20, delay='2')
                                                str_lst = results.split(',')
                                                float_lst = [float(x) for x in str_lst]

                                            # Verify Vout
                                            try:
                                                kvout = float_lst[eval(ConfigMgr.testConditions101['kvout']) - 1]
                                            except:
                                                kvout = 0
                                            if kvout < 0.4:
                                                self.load.output('OFF')
                                                cont = messagebox.askyesno('VOUT NOT DETECTED', 'NO VOUT. Should the test proceed anyway?\nStoping here will not generate the final report.')
                                                if not cont:
                                                    return
                                                else:
                                                    if self.dongle is not None:
                                                        self.dongle.set_frequency(f)
                                                        sleep(0.5)
                                                        self.dongle.set_vout(vout)
                                                        sleep(0.5)
                                                    self.load.output('ON')
                                                    sleep(eval(ConfigMgr.testConditions101['soak']) - 2)

                                                # keithley reading.
                                                if self.keith is not None:
                                                    results = self.keith.read(num=20, delay='2')
                                                    str_lst = results.split(',')
                                                    float_lst = [float(x) for x in str_lst]


                                            # placing data on variables
                                            try:
                                                kiout = float_lst[eval(ConfigMgr.testConditions101['kiout']) - 1]
                                            except:
                                                kiout = 0
                                            try:
                                                iload = 1000 * kiout / eval(ConfigMgr.testConditions101['kioutr'])
                                            except:
                                                iload = 0
                                            try:
                                                kvin = float_lst[eval(ConfigMgr.testConditions101['kvin']) - 1]
                                            except:
                                                kvin = 0
                                            try:
                                                tmon = float_lst[eval(ConfigMgr.testConditions101['ktmon']) - 1]
                                            except:
                                                tmon = 0
                                            try:
                                                kvout = float_lst[eval(ConfigMgr.testConditions101['kvout']) - 1]
                                            except:
                                                kvout = 0
                                            try:
                                                imon = float_lst[eval(ConfigMgr.testConditions101['kimon']) - 1] * 1000
                                            except:
                                                imon = 0
                                            try:
                                                siin = float_lst[eval(ConfigMgr.testConditions101['kiin']) - 1]
                                            except:
                                                siin = 0
                                            try:
                                                iiin = 1000 * siin / eval(ConfigMgr.testConditions101['kiinr'])
                                            except:
                                                iiin = 0
                                            try:
                                                kvcc = float_lst[eval(ConfigMgr.testConditions101['kvcc']) - 1]
                                            except:
                                                kvcc = 0
                                            try:
                                                sicc = float_lst[eval(ConfigMgr.testConditions101['kicc']) - 1]
                                            except:
                                                sicc = 0
                                            try:
                                                iicc = 1000 * sicc / eval(ConfigMgr.testConditions101['kiccr'])
                                            except:
                                                iicc = 0
                                            try:
                                                kven = float_lst[eval(ConfigMgr.testConditions101['kven']) - 1]
                                            except:
                                                kven = 0
                                            try:
                                                sien = float_lst[eval(ConfigMgr.testConditions101['kien']) - 1]
                                            except:
                                                sien = 0
                                            try:
                                                iien = 1000 * sien / eval(ConfigMgr.testConditions101['kienr'])
                                            except:
                                                iien = 0
                                            try:
                                                kvoutsw = float_lst[eval(ConfigMgr.testConditions101['kvoutsw']) - 1]
                                            except:
                                                kvoutsw = 0

                                            pvcc = kvcc * iicc
                                            pin = kvin * iiin
                                            pout = kvout * iload
                                            try:
                                                effeciency = (pout / (pvcc + pin)) * 100
                                            except:
                                                effeciency = 0
                                            iout = iload
                                            try:
                                                temp = self.dongle.get_temp()
                                            except:
                                                temp = 0

                                            # Table update info
                                            if tableID % 2 == 0:
                                                self.table.insert(
                                                    parent='',
                                                    index='end',
                                                    iid=tableID,
                                                    text='',
                                                    values=(mydate, mytime, kvin, iiin, kvcc, iicc, kvout, iload, effeciency),
                                                    tags=('evenrow',)
                                                )
                                            else:
                                                self.table.insert(
                                                    parent='',
                                                    index='end',
                                                    iid=tableID,
                                                    text='',
                                                    values=(mydate, mytime, kvin, iiin, kvcc, iicc, kvout, iload, effeciency),
                                                    tags=('oddrow',)
                                                )
                                            tableID += 1

                                            self.yAxis.append(float(effeciency))
                                            self.xAxis.append(float(i))
                                            self.a.set_ylim(70, 100)
                                            self.a.plot(self.xAxis, self.yAxis)
                                            self.plotCanvas.draw_idle()

                                            dataList.append(
                                                {
                                                    f'{i}A': {
                                                        'vcc_step': vcc,
                                                        'vin_step': vin,
                                                        'vout_step': vout,
                                                        'freq_step': f,
                                                        'kiout': kiout,
                                                        'kvout': kvout,
                                                        'iload': iload,
                                                        'kvin': kvin,
                                                        'siin': siin,
                                                        'iiin': iiin,
                                                        'kvcc': kvcc,
                                                        'sicc': sicc,
                                                        'iicc': iicc,
                                                        'tmon': tmon,
                                                        'kvoutsw': kvoutsw,
                                                        'pin': pin,
                                                        'pvcc': pvcc,
                                                        'pout': pout,
                                                        'effeciency': effeciency,
                                                        'temp': temp,
                                                        'imon': imon,
                                                        'deadRise': deadRise,
                                                        'deadFall': deadFall,
                                                        'sw_max': sw_max,
                                                        'sw_rise': sw_rise,
                                                        'sw_fall': sw_fall,
                                                        'gatel_rise': gatel_rise,
                                                        'gatel_fall': gatel_fall,
                                                    }
                                                }
                                            )

                                            # Store the data in the Excel sheet (in memory until sheet is saved)...
                                            ws1.cell(column=4, row=RowIndex, value=vcc)
                                            ws1.cell(column=5, row=RowIndex, value=mydate)
                                            ws1.cell(column=6, row=RowIndex, value=mytime)
                                            ws1.cell(column=7, row=RowIndex, value=vin)
                                            ws1.cell(column=8, row=RowIndex, value=vout)
                                            ws1.cell(column=9, row=RowIndex, value=f)
                                            ws1.cell(column=10, row=RowIndex, value=i)

                                            ws1.cell(column=12, row=RowIndex, value=kvin)
                                            ws1.cell(column=13, row=RowIndex, value=(siin * 1000))
                                            ws1.cell(column=14, row=RowIndex, value=f'=(M{RowIndex} / B11)')
                                            ws1.cell(column=15, row=RowIndex, value=kvcc)
                                            ws1.cell(column=16, row=RowIndex, value=(sicc * 1000))
                                            ws1.cell(column=17, row=RowIndex, value=f"=(P{RowIndex} / B12)")
                                            ws1.cell(column=18, row=RowIndex, value=kven)
                                            ws1.cell(column=19, row=RowIndex, value=(sien * 1000))
                                            ws1.cell(column=20, row=RowIndex, value=f"=(S{RowIndex} / B13)")
                                            ws1.cell(column=21, row=RowIndex, value=kvout)
                                            ws1.cell(column=22, row=RowIndex, value=(kiout * 1000))
                                            ws1.cell(column=23, row=RowIndex, value=f"=(V{RowIndex} / B14)")
                                            ws1.cell(column=24, row=RowIndex, value=tmon)
                                            ws1.cell(column=25, row=RowIndex, value=kvoutsw)
                                            ws1.cell(column=26, row=RowIndex, value=f"=(L{RowIndex} * N{RowIndex})")
                                            ws1.cell(column=27, row=RowIndex, value=f"=(O{RowIndex} * Q{RowIndex})")
                                            ws1.cell(column=28, row=RowIndex, value=f"=(R{RowIndex} * T{RowIndex})")
                                            ws1.cell(column=29, row=RowIndex, value=f"=(U{RowIndex} * W{RowIndex})")
                                            ws1.cell(column=30, row=RowIndex, value=f"=((Z{RowIndex} + AA{RowIndex} + AB{RowIndex}) - AC{RowIndex})")
                                            ws1.cell(column=31, row=RowIndex, value=f"=(100 * AC{RowIndex}/(Z{RowIndex} + AA{RowIndex} + AB{RowIndex}))")
                                            ws1.cell(column=32, row=RowIndex, value=f"=(X{RowIndex} - 0.6) / 0.008")
                                            ws1.cell(column=33, row=RowIndex, value=imon)
                                            ws1.cell(column=34, row=RowIndex, value=(deadRise * 1e+9))
                                            ws1.cell(column=35, row=RowIndex, value=(deadFall * 1e+9))

                                            # Advance to the next data row...
                                            RowIndex = RowIndex + 1
                                            fileName = folderName + '/data/efficiency.xlsx'
                                            wb.save(filename=fileName)

                                            with open(folderName + f'/data/efficiency.csv', '+a') as csvFile:
                                                csvFile.write(f'{mydate},{mytime},{t}C_{EN}EN_{vcc}Vcc_{f}kHz_{vout}Vout_{vin}Vin,{i},{kvin},{(siin * 1000)},{iiin},{kvcc},{(sicc * 1000)},{iicc},{kven},{sien * 1000},{iien},{kvout},{(kiout * 1000)},{iload},{tmon},{kvoutsw},{pin},{pvcc},{pout},{(pin + pvcc) - pout},{effeciency},{temp},{imon},{(deadRise * 1e+9)},{(deadFall * 1e+9)}\n')

                                        # Scope
                                        if self.scope is not None:
                                            self.scope.run()
                                        if self.load is not None:
                                            self.load.set_current('STATIC', 0)
                                            self.load.output('OFF')
                                            loadOn = 0

                                        dataVin.append({f'{vin}Vin': dataList})
                                    self.f.savefig(folderName + f'/images/efficiency_graph_{f}kHz')
                                    self.xAxis = []
                                    self.yAxis = []
                                    self.a.clear()
                                    self.a.set_ylim(70,100)
                                    self.a.plot(self.xAxis, self.yAxis)
                                    self.plotCanvas.draw_idle()

                                    if eval(ConfigMgr.testConditions101['run2']):
                                        self.statusLabel['text'] = 'Starting get sw slope phase...'
                                        if self.scope is not None:
                                            self.scope.horpos(50)
                                            self.scope.set_vertical_scale(ch=swCh, value=eval(ConfigMgr.instr[f'scopeVertScaleCh{swCh}']))
                                        for ii in self.currList2:
                                            self.load.set_current('STATIC', ii)
                                            if ii:
                                                self.load.output('ON')
                                            else:
                                                if self.scope is not None:
                                                    self.scope.horscale(1/(f*1000))
                                                    self.scope.saveimage(folderName + f'/images/frequency_{t}C_{vcc}Vcc_{vout}Vout_{vin}Vin_{f}kHz_{ii}A.png')

                                            if self.scope is not None:
                                                self.scope.vertical_cursor_off()
                                                self.scope.set_screen_text(text=f"{inductorVal}nH, {f}kHz, {vcc}Vcc, {vout}Vout, {vin}Vin, {ii}A", text_no=1, xpos="1", ypos="8", FontSize=12)
                                                self.scope.set_vertical_scale(swCh, eval(ConfigMgr.instr[f'scopeVertScaleCh{swCh}']))
                                                self.scope.verpos(swCh, eval(ConfigMgr.instr[f'scopeVertPosCh{swCh}']))
                                                self.scope.run()
                                                self.scope.horscale(ConfigMgr.instr['scopeHorScale'])
                                                self.scope.set_vertical_scale(swCh, eval(ConfigMgr.instr[f'scopeVertScaleCh{swCh}']))
                                                self.scope.verpos(swCh, eval(ConfigMgr.instr[f'scopeVertPosCh{swCh}']))
                                                self.scope.Trigger_Edge(
                                                    Level=2.0,
                                                    Slope='RISE',
                                                    Position=2,
                                                    Coupling='DC',
                                                    ChannelIndex=int(swCh)
                                                )
                                                self.scope.trigger_mode('AUTO')
                                                self.scope.set_measurement(
                                                    MeasureIndex=1,
                                                    MeasureType='RISETIME',
                                                    State='ON',
                                                    Source=swCh,
                                                )
                                                if gateLCh:
                                                    self.scope.set_measurement(
                                                        MeasureIndex=2,
                                                        MeasureType='FALLTIME',
                                                        State='ON',
                                                        Source=gateLCh,
                                                    )
                                                self.scope.set_measurement(
                                                    MeasureIndex=3,
                                                    MeasureType='MAXIMUM',
                                                    State='ON',
                                                    Source=swCh,
                                                )
                                                sleep(eval(ConfigMgr.testConditions101['soak']))
                                                self.scope.stop()
                                                try:
                                                    wave = self.scopeSwCh.GetProbeWaveform(Timeout=0.5)
                                                    if self.scope.ID in ['4k', 'TekScope']:
                                                        slicedWave = wave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                        pointA = slicedWave.Measurements_Utils.find_n_maxima(min_find_count=2)
                                                        points = list()
                                                        for p in pointA:
                                                            points.append(p)
                                                    else:
                                                        pointA = wave.Measurements_Utils.find_n_maxima(min_find_count=2)
                                                        points = list()
                                                        for p in pointA:
                                                            points.append(p - (wave.index_to_time(-1) * 0.5))
                                                except:
                                                    points = [0,0]

                                                pointA = points[0]
                                                pointB = points[1]
                                                if not pointA:
                                                    pointA = 0
                                                if not pointB:
                                                    pointB = 0

                                                fileName_i = folderName + f'/images/SW_rise_{t}C_{vcc}Vcc_{vout}Vout_{vin}Vin_{f}kHz_{ii}A.png'
                                                self.scope.saveimage(fileName_i, Background=BACKGROUND)

                                                sw_ring_period = float(self.scope.get_delta_t())
                                                if sw_ring_period == 0:
                                                    sw_ring_freq = 0
                                                else:
                                                    sw_ring_freq = 1 / sw_ring_period * 1e-6
                                                sw_rise = self.scope.get_mean_measurement(Index='1')
                                                if gateLCh:
                                                    gatel_fall = self.scope.get_mean_measurement(Index='2')
                                                else:
                                                    gatel_fall = "0"
                                                sw_max = self.scope.get_mean_measurement(Index='3')
                                                # CDVDT
                                                maxGLCDVDT = 0
                                                if ii > 2 and gateLCh:
                                                    try:
                                                        waveGL = self.scopeGateLCh.GetProbeWaveform(Timeout=0.5)
                                                        if self.scope.ID in ['4k', 'TekScope']:
                                                            slicedWave = waveGL.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                                            pointA = slicedWave.Measurements_Utils.find_n_maxima(min_find_count=2, return_indices=True)
                                                            points = list()
                                                            for p in pointA:
                                                                points.append(p)
                                                            maxGLCDVDT = slicedWave.data[points[1]]
                                                        else:
                                                            pointA = waveGL.Measurements_Utils.find_n_maxima(min_find_count=2, return_indices=True)
                                                            points = list()
                                                            for p in pointA:
                                                                points.append(p)
                                                            maxGLCDVDT = waveGL.data[points[1]]
                                                    except:
                                                        maxGLCDVDT = 0
                                                self.scope.run()

                                                self.scope.Trigger_Edge(
                                                    Level=2.0,
                                                    Slope='FALL',
                                                    Position=2,
                                                    Coupling='DC',
                                                    ChannelIndex=int(swCh)
                                                )
                                                self.scope.trigger_mode('AUTO')
                                                self.scope.set_measurement(
                                                    MeasureIndex=1,
                                                    MeasureType='FALLTIME',
                                                    State='ON',
                                                    Source=swCh,
                                                )
                                                if gateLCh:
                                                    self.scope.set_measurement(
                                                        MeasureIndex=2,
                                                        MeasureType='RISETIME',
                                                        State='ON',
                                                        Source=gateLCh,
                                                    )
                                                if vdshCh:
                                                    self.scope.set_measurement(
                                                        MeasureIndex=3,
                                                        MeasureType='MAXIMUM',
                                                        State='ON',
                                                        Source=vdshCh,
                                                    )
                                                self.scope.vertical_cursor_off()
                                                sleep(eval(ConfigMgr.testConditions101['soak']))
                                                self.scope.stop()
                                                fileName_i = folderName + f'/images/SW_fall_{t}C_{vcc}Vcc_{vout}Vout_{vin}Vin_{f}kHz_{ii}A.png'
                                                self.scope.saveimage(fileName_i, Background=BACKGROUND)
                                                sw_fall = self.scope.get_mean_measurement(Index='1')
                                                if gateLCh:
                                                    gatel_rise = self.scope.get_mean_measurement(Index='2')
                                                else:
                                                    gatel_rise = "0"
                                                vdshMax = 0
                                                if vdshCh:
                                                    vdshMax = self.scope.get_mean_measurement(Index='3')
                                                self.scope.run()


                                            ws1.cell(column=37, row=RowIndex2, value=ii)
                                            ws1.cell(column=38, row=RowIndex2, value=sw_max)
                                            try:
                                                ws1.cell(column=39, row=RowIndex2, value=(eval(sw_rise) * 1e+9))
                                            except:
                                                ws1.cell(column=39, row=RowIndex2, value=0)
                                            try:
                                                ws1.cell(column=40, row=RowIndex2, value=(eval(sw_fall) * 1e+9))
                                            except:
                                                ws1.cell(column=40, row=RowIndex2, value=0)
                                            try:
                                                ws1.cell(column=41, row=RowIndex2, value=(eval(gatel_rise) * 1e+9))
                                            except:
                                                ws1.cell(column=41, row=RowIndex2, value=0)
                                            try:
                                                ws1.cell(column=42, row=RowIndex2, value=(eval(gatel_fall) * 1e+9))
                                            except:
                                                ws1.cell(column=42, row=RowIndex2, value=0)
                                            try:
                                                ws1.cell(column=43, row=RowIndex2, value=vdshMax)
                                            except:
                                                ws1.cell(column=43, row=RowIndex2, value=0)
                                            try:
                                                ws1.cell(column=44, row=RowIndex2, value=sw_ring_freq)
                                            except:
                                                ws1.cell(column=44, row=RowIndex2, value=0)
                                            try:
                                                ws1.cell(column=45, row=RowIndex2, value=maxGLCDVDT)
                                            except:
                                                ws1.cell(column=45, row=RowIndex2, value=0)

                                            RowIndex2 += 1

                                            with open(folderName + f'/data/switch.csv', '+a') as csvFile:
                                                csvFile.write(f',,{t}C_{EN}EN_{vcc}Vcc_{f}kHz_{vout}Vout_{vin}Vin,{ii},{eval(sw_max)},{(eval(sw_rise) * 1e+9)},{(eval(sw_fall) * 1e+9)},{(eval(gatel_rise) * 1e+9)},{(eval(gatel_fall) * 1e+9)},{float(vdshMax)},{sw_ring_freq},{maxGLCDVDT}\n')


                                            try:
                                                dataList.append(
                                                    {
                                                        f'{ii}A': {
                                                            'vdshMax': vdshMax,
                                                            'swRingFreq': sw_ring_freq,
                                                            'maxGLCDVDT': maxGLCDVDT,
                                                        }
                                                    }
                                                )
                                            except:
                                                dataList.append(
                                                    {
                                                        f'{ii}A': {
                                                            'vdshMax': 0,
                                                            'swRingFreq': 0,
                                                            'maxGLCDVDT': 0,
                                                        }
                                                    }
                                                )
                                            fileName = folderName + '/data/efficiency.xlsx'
                                            wb.save(filename=fileName)

                                        if self.scope is not None:
                                            self.scope.run()
                                        if self.load is not None:
                                            self.load.set_current('STATIC', 0)
                                            self.load.output('OFF')
                                        dataVin.append({f'{vin}Vin': dataList})

                                    dataF.append({f'{f}kHz': dataVin})

                                    if eval(ConfigMgr.testConditions101['run3']):
                                        self.statusLabel['text'] = 'Starting snapshot...'
                                        for ii in self.currList2:
                                            self.statusLabel['text'] = f'Running Condition\n{t}C_{EN}Ven_{vcc}Vcc_{f}kHz_{vout}Vout_{vin}Vin_{ii}A'
                                            self.load.set_current('STATIC', ii)
                                            if ii:
                                                self.load.output('ON')

                                            if eval(ConfigMgr.testConditions101['snapPause']):
                                                messagebox.showinfo('first Snap', f'Adjust scope for first screenshot at\n{t}C_{EN}Ven_{vcc}Vcc_{f}kHz_{vout}Vout_{vin}Vin_{ii}A')
                                            else:
                                                sleep(eval(ConfigMgr.testConditions101['soak']))

                                            if self.scope is not None:
                                                self.scope.stop()
                                                sleep(0.5)
                                                fileName_i = folderName + f'/images/snapshot_{t}C_{EN}Ven_{vcc}Vcc_{f}kHz_{vout}Vout_{vin}Vin_{ii}A.png'
                                                self.scope.saveimage(fileName_i, Background=BACKGROUND)
                                                sleep(0.5)

                                                if eval(ConfigMgr.testConditions101['snapScopeZoom']):
                                                    messagebox.showwarning('Second Shot', f'Adjust Scope for second screenshot at\n{t}C_{EN}Ven_{vcc}Vcc_{f}kHz_{vout}Vout_{vin}Vin_{ii}A')
                                                    fileName_i = folderName + f'/images/snapshot_zoom_{t}C_{EN}Ven_{vcc}Vcc_{f}kHz_{vout}Vout_{vin}Vin_{ii}A.png'
                                                    sleep(0.5)
                                                    self.scope.saveimage(fileName_i, Background=BACKGROUND)
                                                self.scope.run()


                                            try:
                                                dataList.append(
                                                    {
                                                        f'{ii}A': {
                                                        }
                                                    }
                                                )
                                            except:
                                                dataList.append(
                                                    {
                                                        f'{ii}A': {
                                                        }
                                                    }
                                                )

                                            if eval(ConfigMgr.testConditions101['loadRest']):
                                                if self.load is not None:
                                                    self.load.output('OFF')
                                                    sleep(eval(ConfigMgr.testConditions101['loadRestTime']))

                                        dataF.append({f'{f}kHz': dataVin})
                            dataVcc.append({f'{vcc}Vcc': dataF})
                        dataVout.append({f'{vout}Vout': dataVcc})
                    test101Dict[f'{t}C'] = dataVout

                    # Save the output file...
                    fileName = folderName + '/data/efficiency.xlsx'
                    wb.save(filename=fileName)

                    with open(folderName + '/data/efficiency_test.json', 'w') as f:
                        json.dump(test101Dict, f, indent=4, ensure_ascii=False)

                    if self.load is not None:
                        self.load.set_current('STATIC', 0)
                        self.load.output('OFF')
            if self.vin is not None:
                self.vin.ch.Enable(False)
            if self.vcc is not None:
                self.vcc.vdrvCh.Enable(False)
            if self.en is not None:
                self.en.ch.Enable(False)
            if self.dongle is not None:
                self.dongle.close()
                self.dongle = None
            if self.scope is not None:
                self.scope = None
                self.scopePwmCh = None
                self.scopeSwCh = None
                self.scopeGateLCh = None
                self.scopeVdshCh = None
            self.statusLabel['text'] = 'Generating Reports...'
            if eval(ConfigMgr.testConditions101['run1']):
                self.run.destFolder = folderName + '/data'
                self.run.finalReport = 'eff_finalReport.xlsx'
                self.run.filePaths = [folderName + '/data/efficiency.csv']
                self.run.generic_report_gen(test='eff')
            if eval(ConfigMgr.testConditions101['run2']):
                self.run.destFolder = folderName + '/data'
                self.run.finalReport = 'sw_finalReport.xlsx'
                self.run.filePaths = [folderName + '/data/switch.csv']
                self.run.generic_report_gen(test='sw')
            messagebox.showinfo('Test Done!', 'Test done!')
        except Exception as e:
            print(e)
            if self.load is not None:
                self.load.set_current('STATIC', 0)
                self.load.output('OFF')
    #endregion

    # region MAKE FOLDER
    def make_dir(self, name: str):
        date = dt.datetime.now()
        date = date.strftime("%Y_%m_%d_%H_%M_%S")
        test = name + date
        isDataDir = os.path.isdir('Data')
        if (not isDataDir):
            os.mkdir('Data')
        isTestDir = os.path.isdir('Data/' + str(test))
        if (not isTestDir):
            os.mkdir('Data/' + str(test))
        isTestImgDir = os.path.isdir('Data/' + str(test) + '/images')
        if (not isTestImgDir):
            os.mkdir('Data/' + str(test) + '/images')
        isTestDataDir = os.path.isdir('Data/' + str(test) + '/data')
        if (not isTestDataDir):
            os.mkdir('Data/' + str(test) + '/data')
        return 'Data/' + test
    # endregion

    def create_excel(self, title='', index=0, config=None):
        # Create Excel File
        wb = Workbook()
        ws1 = wb.create_sheet(title=title, index=index)

        # Set up Excel Doc:
        # Parameter ...

        ws1.column_dimensions['A'].width = 24
        ws1.column_dimensions['B'].width = 9
        r = 2
        ws1.cell(column=1, row=r, value="Start Current (A):")
        ws1.cell(column=2, row=r, value=config['startCurr'])
        r += 1
        ws1.cell(column=1, row=r, value="Final Current(A):")
        ws1.cell(column=2, row=r, value=config['endCurr'])
        r += 1
        ws1.cell(column=1, row=r, value="Increments (A):")
        ws1.cell(column=2, row=r, value=config['stepCurr'])
        r += 1
        ws1.cell(column=1, row=r, value="Time for statistics count (s):")
        ws1.cell(column=2, row=r, value=eval(config['soak']))
        r += 1
        ws1.cell(column=1, row=r, value="Vdrv list (V):")
        ws1.cell(column=2, row=r, value=str(self.vccList))
        r += 1
        ws1.cell(column=1, row=r, value="Vin list (V):")
        ws1.cell(column=2, row=r, value=str(self.pvinList))
        r += 1
        ws1.cell(column=1, row=r, value="Fsw list (kHz):")
        ws1.cell(column=2, row=r, value=str(self.freqList))
        r += 1
        ws1.cell(column=1, row=r, value="Vout list (V):")
        ws1.cell(column=2, row=r, value=str(self.voutList))
        r += 1
        ws1.cell(column=1, row=r, value="Inductor (nH):")
        ws1.cell(column=2, row=r, value=str(config['inductorVal']))
        r += 1
        ws1.cell(column=1, row=r, value="Iin Shunt (m\u03A9):")
        ws1.cell(column=2, row=r, value=float(config['kiinr']))
        r += 1
        ws1.cell(column=1, row=r, value="Icc Shunt (m\u03A9):")
        ws1.cell(column=2, row=r, value=float(config['kiccr']))
        r += 1
        ws1.cell(column=1, row=r, value="Ien Shunt (m\u03A9):")
        ws1.cell(column=2, row=r, value=float(config['kienr']))
        r += 1
        ws1.cell(column=1, row=r, value="Iout Shunt (m\u03A9):")
        ws1.cell(column=2, row=r, value=float(config['kioutr']))
        # data
        if title == 'Vout Ripple Fsw vs Load':
            self.data104_excel(ws1)
        if title == 'Efficiency':
            self.data101_excel(ws1)
        return wb

    def data101_excel(self, ws1):
        c = 4
        ws1.cell(column=c, row=2, value='VCC')
        c += 1
        ws1.cell(column=c, row=2, value='Date')
        c += 1
        ws1.cell(column=c, row=2, value='Time')
        c += 1
        ws1.cell(column=c, row=2, value='Pvin (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Vout (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Frequency (kHz)')
        c += 1
        ws1.cell(column=c, row=2, value='Load (A)')
        c += 2
        ws1.cell(column=c, row=2, value='Vin (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Iin (mV)')
        c += 1
        ws1.cell(column=c, row=2, value='Iin (A)')
        c += 1
        ws1.cell(column=c, row=2, value='Vcc (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Icc (mV)')
        c += 1
        ws1.cell(column=c, row=2, value='Icc (A)')
        c += 1
        ws1.cell(column=c, row=2, value='EN (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Ien (mV)')
        c += 1
        ws1.cell(column=c, row=2, value='Ien (A)')
        c += 1
        ws1.cell(column=c, row=2, value='Vout (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Iout (mV)')
        c += 1
        ws1.cell(column=c, row=2, value='Iout (A)')
        c += 1
        ws1.cell(column=c, row=2, value='Tmon (V)')
        c += 1
        ws1.cell(column=c, row=2, value='VoutSW (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Pin (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Pvcc (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Pven (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Pout (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Power Loss (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Efficiency (%)')
        c += 1
        ws1.cell(column=c, row=2, value='Tmon (C)')
        c += 1
        ws1.cell(column=c, row=2, value='Imon (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Dead Time Rise (ns)')
        c += 1
        ws1.cell(column=c, row=2, value='Dead Time Fall (ns)')
        c += 1
        # empty column
        c += 1
        ws1.cell(column=c, row=2, value='Load (A)')
        c += 1
        ws1.cell(column=c, row=2, value='SW Max (V)')
        c += 1
        ws1.cell(column=c, row=2, value='SW Rise Time (ns)')
        c += 1
        ws1.cell(column=c, row=2, value='SW Fall Time (ns)')
        c += 1
        ws1.cell(column=c, row=2, value='GateL Rise Time (ns)')
        c += 1
        ws1.cell(column=c, row=2, value='GateL Fall Time (ns)')
        c += 1
        ws1.cell(column=c, row=2, value='VDSH Max (V)')
        c += 1
        ws1.cell(column=c, row=2, value='sw ring Freq (MHz)')
        c += 1
        ws1.cell(column=c, row=2, value='Max GL CdVdt (V)')

    def init_scope(self):
        sc = None
        if eval(ConfigMgr.instr['scopeOnOff']):
            scopeModel = ConfigMgr.instr['scopeModel']
            addr = ConfigMgr.instr['scopeAddr']
            if scopeModel in ['TDS5104B', 'MSO5204B', 'MSO5204']:
                sc = purescope.TekScope(addr, Simulate=False, Reset=False)
            elif (scopeModel in ['MSO56', 'MSO54', 'MSO58', 'MSO58LP']): # THIS DRIVER IS NOT WORKING PROPERLY FOR NOW
                sc = purescope.TEK_MSO5X(addr, Simulate=False, Reset=False)
            elif (scopeModel in ['MSO56', 'MSO54', 'MSO58', 'MSO58LP', 'MSO58B']):
                sc = purescope.TEK_MSO5XB(addr, Simulate=False, Reset=False)
            elif (scopeModel in ['MSO4104', 'MDO3104']):
                sc = purescope.Tkdpo4k(addr, Simulate=False, Reset=False)
            elif (scopeModel in ['DPO7104', 'DPO7104C']):
                sc = purescope.Tkdpo7k(addr, Simulate=False, Reset=False)
            else:
                print('Oscilloscope not defined.')
        return sc
