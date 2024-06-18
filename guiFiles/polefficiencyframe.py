import tkinter as tk
from tkinter import ttk, StringVar, messagebox
import guiFiles.guihelperfuncpol as make
import guiFiles.guivalidatefunc as validate
import guiFiles.initinstruments as init
import guiFiles.auxfunclib as aux
from guiFiles.tk_tooltip import CreateToolTip
from guiFiles.configmgr import ConfigMgr
from guiFiles.killthread import KillableThread

import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.animation as animation
from matplotlib import style
style.use('ggplot')

from time import sleep
import json
import os
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Alignment

class EffFrame(tk.Frame):

    tempList = None
    currList = None
    freqList = None
    pvinList = None
    voutList = None
    vccList = None

    load = None
    keith = None
    vcc = None
    vin = None
    scope = None

    def __init__(self, parent):
        super(EffFrame, self).__init__(parent)
        self.config(background='white')
        self.r = 0
        self.c = 0
        self.style = ttk.Style()
        # self.style.theme_use('default')
        self.style.configure(
            'Treeview',
            background='#D3D3D3',
            foreground='black',
            rowheight=25,
            fieldbackground='#D3D3D3'
        )
        self.style.map(
            'Treeview',
            background=[('selected', '#347083')],
        )

        # region JIRA
        self.jira = tk.LabelFrame(self, text='JIRA', background='white')
        self.jira.grid(row=self.r, column=self.c, padx=10, sticky='w')
        self.build_box_jira()
        self.c+=1
        #endregiom

        # region Chamber Temperatue
        self.Temp = tk.LabelFrame(self, text='Temperature Steps (\u00b0C)', background='white')
        self.Temp.grid(row=self.r, column=self.c, padx=10, sticky='w')
        self.build_box_temp()
        self.c += 1
        # endregion

        # region Status Bar
        self.statusBar = tk.LabelFrame(self, text='Status', background='white')
        self.statusBar.grid(row=self.r, column=self.c, padx=10, sticky='nsew')
        self.statusLabel = tk.Label(
            self.statusBar,
            text='WELCOME!',
            anchor=tk.CENTER,
            foreground='#0FFF50',
            background='black',
            font=('Arial Bold', 12)
        )
        self.statusLabel.pack(expand=1, fill='both')
        self.r += 1
        self.c -= 1
        # endregion

        # Setup box
        self.setupLblFrame = tk.LabelFrame(self, text='Setup', background='#A4E1FF')
        self.setupLblFrame.grid(row=self.r, column=self.c, padx=10, pady=5, sticky='nw')
        self.setup_box()
        self.c += 1

        # Graph
        self.graphFrame = ttk.Frame(self)
        self.graphFrame.grid(row=self.r, column=self.c, rowspan=30, padx=10, pady=5, sticky='nw')
        self.xAxis = list()
        self.yAxis = list()
        self.f = Figure(figsize=(5, 5), dpi=100)
        self.a = self.f.add_subplot(111)
        self.a.set_ylim(0, 100)
        self.graphF = ttk.Frame(self.graphFrame)
        self.graphF.grid(row=0, column=0)
        self.plotCanvas = FigureCanvasTkAgg(self.f, master=self.graphF)
        self.plotCanvas.get_tk_widget().grid(row=0, column=0, padx=2, pady=2)
        # self.graph_box()
        self.r += 1
        self.c -= 1

        # Table
        self.tableFrame = tk.Frame(self)
        self.tableFrame.grid(row=self.r, column=self.c, padx=10, rowspan=28, sticky='nsew')
        self.tableFrame.grid_propagate(0)
        tableScrollY = tk.Scrollbar(self.tableFrame)
        tableScrollY.pack(side=tk.RIGHT, fill=tk.Y)
        tableScrollX = tk.Scrollbar(self.tableFrame, orient='horizontal')
        tableScrollX.pack(side=tk.BOTTOM, fill=tk.X)
        self.table = ttk.Treeview(
            self.tableFrame,
            yscrollcommand=tableScrollY.set,
            xscrollcommand=tableScrollX.set,
            selectmode='extended',
        )
        self.table.pack(fill=tk.BOTH, expand=True)
        tableScrollY.config(command=self.table.yview)
        tableScrollX.config(command=self.table.xview)
        self.table_box()
        self.r += 28

        # region BUTTONS
        BOLDFONT = tk.font.Font(weight='bold')
        btnFrame = ttk.Frame(self)
        btnFrame.grid(row=self.r, column=self.c, sticky='se', padx=5)
        # region KILL TEST BUTTON
        self.killBtn = tk.Button(
            btnFrame,
            text='KILL TEST',
            padx=5,
            font=BOLDFONT,
            background='red',
            foreground='white',
            command=self.kill_test,
        )
        self.killBtn.grid(row=0, column=0, padx=5)
        CreateToolTip(self.killBtn, 'Most or ALL DATA will be LOST!')
        self.killBtn['state'] = 'disabled'
        # endregion

        # region RUN BUTTON
        self.runButton = tk.Button(
            btnFrame,
            text='Run Test',
            padx=5,
            font=BOLDFONT,
            background='#003833',
            foreground='white',
            command=self.run_tests,
        )
        self.runButton.grid(row=0, column=1, padx=5)
        # endregion
        # endregion

    def build_box_temp(self):
        self.tempFrame = tk.Frame(self.Temp, background='#FFB4A4')
        self.tempFrame.grid(row=0, column=0, columnspan=5, rowspan=3)
        self.varTempStep1 = StringVar()
        self.varTempStep1.set(ConfigMgr.tempSteps['1'])
        self.varTempStep2 = StringVar()
        self.varTempStep2.set(ConfigMgr.tempSteps['2'])
        self.varTempStep3 = StringVar()
        self.varTempStep3.set(ConfigMgr.tempSteps['3'])
        self.varTempStep4 = StringVar()
        self.varTempStep4.set(ConfigMgr.tempSteps['4'])
        self.varTempStep5 = StringVar()
        self.varTempStep5.set(ConfigMgr.tempSteps['5'])
        self.varTempStep6 = StringVar()
        self.varTempStep6.set(ConfigMgr.tempSteps['6'])
        tempArgList = [
            self.tempFrame,
            ConfigMgr.tempSteps,
            validate.temperature,
            self.varTempStep1,
            self.varTempStep2,
            self.varTempStep3,
            self.varTempStep4,
            self.varTempStep5,
            self.varTempStep6,
        ]
        make.temperature_box(tempArgList)
    # endregion

    def build_box_jira(self):
        self.tempFrame = tk.Frame(self.Temp, background='#FFB4A4')
        self.tempFrame.grid(row=0, column=0, columnspan=5, rowspan=3)
        make.jira_box()

    def setup_box(self):
        # Vcc Steps
        vccFrame = tk.LabelFrame(self.setupLblFrame, text='Vcc Steps (V)', background='#A4E1FF')
        vccFrame.grid(row=0, column=0, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartVcc = StringVar()
        lbl1 = tk.Label(vccFrame, text='Start Vcc:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            vccFrame,
            textvariable=self.varStartVcc,
            validate='focusout',
            validatecommand=self.validate_vcc,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartVcc.set(ConfigMgr.testConditions101['startVcc'])
        self.varEndVcc = StringVar()
        lbl2 = tk.Label(vccFrame, text='End Vcc:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            vccFrame,
            textvariable=self.varEndVcc,
            validate='focusout',
            validatecommand=self.validate_vcc,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndVcc.set(ConfigMgr.testConditions101['endVcc'])
        self.varStepVcc = StringVar()
        lbl3 = tk.Label(vccFrame, text='Step Vcc:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            vccFrame,
            textvariable=self.varStepVcc,
            validate='focusout',
            validatecommand=self.validate_vcc,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepVcc.set(ConfigMgr.testConditions101['stepVcc'])

        # Vin Steps
        vinFrame = tk.LabelFrame(self.setupLblFrame, text='Vin Steps (V)', background='#A4E1FF')
        vinFrame.grid(row=0, column=1, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartVin = StringVar()
        lbl1 = ttk.Label(vinFrame, text='Start Vin:')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            vinFrame,
            textvariable=self.varStartVin,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartVin.set(ConfigMgr.testConditions101['startPVin'])
        self.varEndVin = StringVar()
        lbl2 = tk.Label(vinFrame, text='End Vin:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            vinFrame,
            textvariable=self.varEndVin,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndVin.set(ConfigMgr.testConditions101['endPVin'])
        self.varStepVin = StringVar()
        lbl3 = tk.Label(vinFrame, text='Step Vin:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            vinFrame,
            textvariable=self.varStepVin,
            validate='focusout',
            validatecommand=self.validate_vin,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepVin.set(ConfigMgr.testConditions101['stepPVin'])

        # Frequency Steps
        fswFrame = tk.LabelFrame(self.setupLblFrame, text='Frequency Steps (kHz)', background='#A4E1FF')
        fswFrame.grid(row=0, column=2, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartFsw = StringVar()
        lbl1 = tk.Label(fswFrame, text='Start Frequency:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            fswFrame,
            textvariable=self.varStartFsw,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartFsw.set(ConfigMgr.testConditions101['startFsw'])
        self.varEndFsw = StringVar()
        lbl2 = tk.Label(fswFrame, text='End Frequency:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            fswFrame,
            textvariable=self.varEndFsw,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndFsw.set(ConfigMgr.testConditions101['endFsw'])
        self.varStepFsw = StringVar()
        lbl3 = tk.Label(fswFrame, text='Step Frequency:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            fswFrame,
            textvariable=self.varStepFsw,
            validate='focusout',
            validatecommand=self.validate_fsw,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepFsw.set(ConfigMgr.testConditions101['stepFsw'])

        # Load Steps
        loadFrame = tk.LabelFrame(self.setupLblFrame, text='Load Steps (A)', background='#A4E1FF')
        loadFrame.grid(row=0, column=3, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartCurrent = StringVar()
        lbl1 = tk.Label(loadFrame, text='Start Current:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            loadFrame,
            textvariable=self.varStartCurrent,
            validate='focusout',
            validatecommand=self.validate_current,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartCurrent.set(ConfigMgr.testConditions101['startCurr'])
        self.varEndCurrent = StringVar()
        lbl2 = tk.Label(loadFrame, text='End Current:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            loadFrame,
            textvariable=self.varEndCurrent,
            validate='focusout',
            validatecommand=self.validate_current,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndCurrent.set(ConfigMgr.testConditions101['endCurr'])
        self.varStepCurrent = StringVar()
        lbl3 = tk.Label(loadFrame, text='Step Current:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            loadFrame,
            textvariable=self.varStepCurrent,
            validate='focusout',
            validatecommand=self.validate_current,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepCurrent.set(ConfigMgr.testConditions101['stepCurr'])

        # Vout Steps
        voutFrame = tk.LabelFrame(self.setupLblFrame, text='Vout Steps (V)', background='#A4E1FF')
        voutFrame.grid(row=0, column=4, rowspan=3, padx=2, pady=5, sticky='nw')
        self.varStartVout = StringVar()
        lbl1 = tk.Label(voutFrame, text='Start Vout:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            voutFrame,
            textvariable=self.varStartVout,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varStartVout.set(ConfigMgr.testConditions101['startVout'])
        self.varEndVout = StringVar()
        lbl2 = tk.Label(voutFrame, text='End Vout:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            voutFrame,
            textvariable=self.varEndVout,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varEndVout.set(ConfigMgr.testConditions101['endVout'])
        self.varStepVout = StringVar()
        lbl3 = tk.Label(voutFrame, text='Step Vout:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            voutFrame,
            textvariable=self.varStepVout,
            validate='focusout',
            validatecommand=self.validate_vout,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varStepVout.set(ConfigMgr.testConditions101['stepVout'])

        # Scope Channels
        scopeFrame = tk.Frame(self.setupLblFrame, background='#A4E1FF')
        scopeFrame.grid(row=0, column=5, padx=2, pady=5, sticky='nw')
        self.varPWM = StringVar()
        lbl1 = tk.Label(scopeFrame, text='PWM Ch:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            scopeFrame,
            textvariable=self.varPWM,
            validate='focusout',
            validatecommand=self.validate_scope,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varPWM.set(ConfigMgr.testConditions101['scopePwmCh'])
        self.varSW = StringVar()
        lbl2 = tk.Label(scopeFrame, text='SW Ch:', background='#A4E1FF')
        lbl2.grid(row=1, column=0, pady=2)
        entry2 = ttk.Entry(
            scopeFrame,
            textvariable=self.varSW,
            validate='focusout',
            validatecommand=self.validate_scope,
            width=4,
        )
        entry2.grid(row=1, column=1, pady=2)
        self.varSW.set(ConfigMgr.testConditions101['scopeSwCh'])
        self.varGateL = StringVar()
        lbl3 = tk.Label(scopeFrame, text='GateL Ch:', background='#A4E1FF')
        lbl3.grid(row=2, column=0, pady=2)
        entry3 = ttk.Entry(
            scopeFrame,
            textvariable=self.varGateL,
            validate='focusout',
            validatecommand=self.validate_scope,
            width=4,
        )
        entry3.grid(row=2, column=1, pady=2)
        self.varGateL.set(ConfigMgr.testConditions101['scopeGateLCh'])

        # Soak Time
        soakFrame = tk.Frame(self.setupLblFrame, background='#A4E1FF')
        soakFrame.grid(row=2, column=5, padx=2, pady=0, sticky='nw')
        self.varSoakTime = StringVar()
        lbl1 = tk.Label(soakFrame, text='Soak Time:', background='#A4E1FF')
        lbl1.grid(row=0, column=0, pady=2)
        entry1 = ttk.Entry(
            soakFrame,
            textvariable=self.varSoakTime,
            validate='focusout',
            validatecommand=self.validate_soak,
            width=4,
        )
        entry1.grid(row=0, column=1, pady=2)
        self.varSoakTime.set(ConfigMgr.testConditions101['soak'])
        lbl1unt = tk.Label(soakFrame, text='s', background='#A4E1FF')
        lbl1unt.grid(row=0, column=2, pady=2)

        # region test 1.01 Keithley Channels
        self.keithChsFrame = tk.LabelFrame(self.setupLblFrame, text='Keithley Channels', background='#A4E1FF')
        self.keithChsFrame.grid(row=3, column=0, columnspan=6, padx=2, pady=(0, 2), sticky='nw')
        self.kiin = StringVar()
        self.kiin.set(ConfigMgr.testConditions101['kiin'])
        self.kimon = StringVar()
        self.kimon.set(ConfigMgr.testConditions101['kimon'])
        self.kvout = StringVar()
        self.kvout.set(ConfigMgr.testConditions101['kvout'])
        self.ktmon = StringVar()
        self.ktmon.set(ConfigMgr.testConditions101['ktmon'])
        self.kvin = StringVar()
        self.kvin.set(ConfigMgr.testConditions101['kvin'])
        self.kiout = StringVar()
        self.kiout.set(ConfigMgr.testConditions101['kiout'])
        self.kioutr = StringVar()
        self.kioutr.set(ConfigMgr.testConditions101['kioutr'])
        self.kiinr = StringVar()
        self.kiinr.set(ConfigMgr.testConditions101['kiinr'])
        self.kvcc = StringVar()
        self.kvcc.set(ConfigMgr.testConditions101['kvcc'])
        self.kicc = StringVar()
        self.kicc.set(ConfigMgr.testConditions101['kicc'])
        self.kiccr = StringVar()
        self.kiccr.set(ConfigMgr.testConditions101['kiccr'])
        self.kvoutsw = StringVar()
        self.kvoutsw.set(ConfigMgr.testConditions101['kvoutsw'])
        self.kChBoxArgList = [
            self.keithChsFrame,
            ConfigMgr.testConditions101,
            validate.keithley_channels,
            self.kvin,
            self.kiin,
            self.kiinr,
            self.kvout,
            self.kiout,
            self.kioutr,
            self.kvcc,
            self.kicc,
            self.kiccr,
            self.ktmon,
            self.kimon,
            self.kvoutsw,
        ]
        make.keithley_channel(self.kChBoxArgList)
        # endregion

    def graph_box(self):
        graphF = ttk.Frame(self.graphFrame)
        graphF.grid(row=0, column=0)

        plotCanvas = FigureCanvasTkAgg(self.f, master=graphF)
        plotCanvas.draw()
        plotCanvas.get_tk_widget().grid(row=0, column=0, padx=2, pady=2)

    def table_box(self):
        self.table['columns'] = (
            'Date',
            'Time',
            'Vin',
            'Iin',
            'Vcc',
            'Icc',
            'Vout',
            'Iout',
            'Efficiency',
            'Temp',
            'Humidity',
        )
        self.table.column('#0', width=0, stretch=tk.NO)
        self.table.column('Date', anchor=tk.W, width=25)
        self.table.column('Time', anchor=tk.W, width=25)
        self.table.column('Vin', anchor=tk.CENTER, width=25)
        self.table.column('Iin', anchor=tk.CENTER, width=25)
        self.table.column('Vcc', anchor=tk.CENTER, width=25)
        self.table.column('Icc', anchor=tk.CENTER, width=25)
        self.table.column('Vout', anchor=tk.CENTER, width=25)
        self.table.column('Iout', anchor=tk.CENTER, width=25)
        self.table.column('Efficiency', anchor=tk.CENTER, width=30)
        self.table.column('Temp', anchor=tk.CENTER, width=25)
        self.table.column('Humidity', anchor=tk.CENTER, width=30)

        self.table.heading('#0', anchor=tk.W, text='')
        self.table.heading('Date', anchor=tk.CENTER, text='Date')
        self.table.heading('Time', anchor=tk.CENTER, text='Time')
        self.table.heading('Vin', anchor=tk.CENTER, text='Vin')
        self.table.heading('Iin', anchor=tk.CENTER, text='Iin')
        self.table.heading('Vcc', anchor=tk.CENTER, text='Vcc')
        self.table.heading('Icc', anchor=tk.CENTER, text='Icc')
        self.table.heading('Vout', anchor=tk.CENTER, text='Vout')
        self.table.heading('Iout', anchor=tk.CENTER, text='Iout')
        self.table.heading('Efficiency', anchor=tk.CENTER, text='Efficiency')
        self.table.heading('Temp', anchor=tk.CENTER, text='Temp.')
        self.table.heading('Humidity', anchor=tk.CENTER, text='Humidity')

        self.table.tag_configure('oddrow', background='white')
        self.table.tag_configure('evenrow', background='lightblue')

    # VALIDATION FUNCTIONS
    def validate_current(self):
        ConfigMgr.testConditions101['startCurr'] = self.varStartCurrent.get()
        ConfigMgr.testConditions101['endCurr'] = self.varEndCurrent.get()
        ConfigMgr.testConditions101['stepCurr'] = self.varStepCurrent.get()
        return True

    def validate_fsw(self):
        ConfigMgr.testConditions101['startFsw'] = self.varStartFsw.get()
        ConfigMgr.testConditions101['endFsw'] = self.varEndFsw.get()
        ConfigMgr.testConditions101['stepFsw'] = self.varStepFsw.get()
        return True

    def validate_vin(self):
        ConfigMgr.testConditions101['startPVin'] = self.varStartVin.get()
        ConfigMgr.testConditions101['endPVin'] = self.varEndVin.get()
        ConfigMgr.testConditions101['stepPVin'] = self.varStepVin.get()
        return True

    def validate_vout(self):
        ConfigMgr.testConditions101['startVout'] = self.varStartVout.get()
        ConfigMgr.testConditions101['endVout'] = self.varEndVout.get()
        ConfigMgr.testConditions101['stepVout'] = self.varStepVout.get()
        return True

    def validate_vcc(self):
        ConfigMgr.testConditions101['startVcc'] = self.varStartVcc.get()
        ConfigMgr.testConditions101['endVcc'] = self.varEndVcc.get()
        ConfigMgr.testConditions101['stepVcc'] = self.varStepVcc.get()
        return True

    def validate_soak(self):
        ConfigMgr.testConditions101['soak'] = self.varSoakTime.get()
        return True

    def validate_scope(self):
        ConfigMgr.testConditions101['scopePwmCh'] = self.varPWM.get()
        ConfigMgr.testConditions101['scopeSwCh'] = self.varSW.get()
        ConfigMgr.testConditions101['scopeGateLCh'] = self.varGateL.get()
        return True

    # RUN TEST
    def run_tests(self):
        self.running = KillableThread(name='test', target=self._run_tests)
        self.running.start()

    def _run_tests(self):
        self.buttons_switch()
        self.run_efficiency()
        self.buttons_switch()
        return

    def kill_test(self):
        self.running.raise_exception()
        self.running.join()
        self.buttons_switch()

    def buttons_switch(self):
        if(str(self.runButton.cget('state')) == 'normal'):
            self.runButton['state'] = 'disabled'
            self.killBtn['state'] = 'normal'
        else:
            self.runButton['state'] = 'normal'
            self.killBtn['state'] = 'disabled'

    def animate_graph(self, interval):
        print('animate_graph called!')
        self.a.clear()
        self.a.plot(self.xAxis, self.yAxis)

    def run_efficiency(self):
        try:
            self.statusLabel['text'] = 'Initializing test Variables...'
            self.setting_pre_test_variables(ConfigMgr.testConditions101)
            self.statusLabel['text'] = 'Creating folder to save data'
            folderName = self.make_dir('efficiency')
            self.statusLabel['text'] = 'Initializing Instruments...'
            if self.load is None:
                self.load = init.eload()
            if self.keith is None:
                self.keith = init.keith()
            if self.vcc is None:
                self.vcc = init.vcc()
            if self.vin is None:
                self.vin = init.vin()
            if self.scope is None:
                self.scope = init.oscilloscope()
            test101Dict = {}
            self.statusLabel['text'] = 'Creating Excel file...'
            wb = self.create_excel(title='Efficiency', index=0, config=ConfigMgr.testConditions101)
            ws1 = wb['Efficiency']
            self.data101_excel(ws1)
            for item in self.table.get_children():
                self.table.delete(item)
            self.xAxis = list()
            self.yAxis = list()
            self.a.clear()
            self.a.plot(self.xAxis, self.yAxis)
            self.plotCanvas.draw_idle()

            # Chroma Set-up
            if self.load is not None:
                self.load.set_channel(index=ConfigMgr.instr['loadChannel'][-1])
                self.load.set_mode(mode='CCH')
                self.load.set_current('STATIC', 0)
                self.load.output('OFF')

            # Keith Set-up
            if self.keith is not None:
                self.keith.reset_keithley()
                self.keith.select_function(function='VOLT:DC', channel='(@101:120)')
                self.keith.select_range(function='VOLT:DC', range='AUTO ON', channel='(@101:120)')
                self.keith.set_nplc(function='VOLT:DC', value='MIN', channel='(@101:120)')
                self.keith.route_to_scan('(@101:120)')
                self.keith.enable_scan('ON')
                self.keith.number_of_channels_to_scan(20)
                self.keith.data_elements(reading=1)

            # VCC Set-up
            if eval(ConfigMgr.instr['vccPSOnOff']):
                if self.vcc is not None:
                    self.vcc.vdrvCh.Configure_VoltageLevel(Level=self.vccList[0], CurrentLimit=1)
                    self.vcc.vdrvCh.Enable(True)

            # VIN Set-up
            if self.vin is not None:
                self.vin.ch.Configure_VoltageLevel(Level=self.pvinList[0], CurrentLimit=20)
                self.vin.ch.Enable(False)

            # SCOPE Set-up
            if self.scope is not None:
                try:
                    ch1 = self.scope.GetChannel(Index=1)
                    ch1.Enable(False)
                except Exception as e:
                    print('Failed Creating Scope channel Object')
                    print(e)
                    return
                try:
                    swCh = eval(ConfigMgr.testConditions101['scopeSwCh'])
                    self.scopeSwCh = self.scope.GetChannel(Index=swCh)
                    self.scopeSwCh.Enable(True)
                    self.scope.set_attenuation(swCh, eval(ConfigMgr.instr[f'scopeAttnCh{swCh}']))
                except:
                    print('Failed Creating SW Scope channel Object')
                    return
                try:
                    pwmCh = eval(ConfigMgr.testConditions101['scopePwmCh'])
                    self.scopePwmCh = self.scope.GetChannel(Index=pwmCh)
                    self.scopePwmCh.Enable(True)
                    self.scope.set_attenuation(pwmCh, eval(ConfigMgr.instr[f'scopeAttnCh{pwmCh}']))
                except:
                    print('Failed Creating PWM Scope channel Object')
                    return
                try:
                    gateLCh = eval(ConfigMgr.testConditions101['scopeGateLCh'])
                    self.scopeGateLCh = self.scope.GetChannel(Index=gateLCh)
                    self.scopeGateLCh.Enable(True)
                    self.scope.set_attenuation(gateLCh, eval(ConfigMgr.instr[f'scopeAttnCh{gateLCh}']))
                except:
                    print('Failed Creating GateL Scope channel Object')
                    return

                # Overlay display
                self.scope.display_overlay()
                self.scope.horizontal_delay('OFF')
                self.scope.set_persistence('OFF')
                self.scope.set_acquisition_mode('HIR')

                # Config Probe Channels
                self.scopeSwCh.ProbeSetup(
                    Coupling="DC",
                    Bandwidth=None,
                    Vrange=int(eval(ConfigMgr.instr[f'scopeVertScaleCh{swCh}']) * 10),
                    Offset=0,
                    Position=0,
                    Impedance=eval(ConfigMgr.instr[f'scopeTermCh{swCh}']),
                    Probe_Attn=eval(ConfigMgr.instr[f'scopeAttnCh{swCh}'])
                )

                # Trigger
                self.scope.Trigger_Edge(
                    Level=2.00,
                    Slope='RISE',
                    Position=2,
                    Coupling='DC',
                    ChannelIndex=int(swCh)
                )

                # Vertical Position
                self.scope.verpos(swCh, str(ConfigMgr.instr['scopeVertPosCh' + str(swCh)]))
                self.scope.set_channel_label(swCh, 'SW', '50', '-30')

                self.scope.verpos(gateLCh, str(ConfigMgr.instr['scopeVertPosCh' + str(gateLCh)]))
                self.scope.set_channel_label(gateLCh, 'GateL', '50', '-30')

                self.scope.verpos(pwmCh, str(ConfigMgr.instr['scopeVertPosCh' + str(pwmCh)]))
                self.scope.set_channel_label(pwmCh, 'PWM', '50', '-30')

                # Horizontal Scale
                self.scope.horscale('4e-9')
                self.scope.horizontal_mode('AUTO')
                self.scope.horizontal_sample_rate('5e+9')
                self.scope.horpos(50)

                # Measurements
                self.scope.set_measurement(
                    MeasureIndex=5,
                    MeasureType='FREQUENCY',
                    State='ON',
                    Source=swCh,
                )

                # Scope Run
                self.scope.Arm(Continuous=True)
                self.scope.set_persistence('1')
                self.scope.vertical_cursor_on()
                self.scope.set_cursor_function('WAVEFORM')
                self.scope.cursor_split()
                self.scope.waveform_cursor_to_channel(cursor=1, ch=swCh)
                self.scope.waveform_cursor_to_channel(cursor=2, ch=swCh)

            # Set the first row for data output...
            RowIndex = 3
            tableID = 0
            sw_freq = 0
            dead1 = 0
            dead2 = 0
            sw_max = 0
            sw_rise = 0
            sw_fall = 0
            gatel_rise = 0
            gatel_fall = 0
            for t in self.tempList:
                if eval(ConfigMgr.instr['thermOnOff']):
                    self.statusLabel['text'] = f'Setting Temperature to {t}\u00b0C'
                    self.therm.set_temperature(t)
                    currTemp = self.chamber_get_temp()
                    currTemp.decode('ascii')
                    currTemp = float(currTemp)
                    while (abs(currTemp - t) > 1):
                        currTemp = self.chamber_get_temp()
                        currTemp.decode('ascii')
                        currTemp = float(currTemp)
                        sleep(60)
                else:
                    messagebox.showwarning('Temperature', f'Set Temperature to {t}\u00B0C')

                dataVout = list()
                for vout in self.voutList:
                    # Changing Vout
                    if ConfigMgr.instr['polBoardModel'] in ['Malibu']:
                        self.statusLabel['text'] = f'Setting Vout to {vout}V'
                        # @TODO code to change Vout
                        pass
                    else:
                        messagebox.showwarning('Vout', f'Set Vout to {vout}V')

                    dataVcc = list()
                    for vcc in self.vccList:
                        if self.vcc is not None:
                            self.statusLabel['text'] = f'Setting Vcc to {vcc}V'
                            self.vcc.vdrvCh.Configure_VoltageLevel(Level=vcc, CurrentLimit=1)
                            self.vcc.vdrvCh.Enable(True)
                        else:
                            messagebox.showwarning('Vcc', f'Set Vcc to {vcc}V')

                        firstFreq = 1
                        dataF = list()
                        for f in self.freqList:
                            # Changing Fsw
                            if ConfigMgr.instr['polBoardModel'] in ['Malibu']:
                                self.statusLabel['text'] = f'Setting Frequency to {f}kHz'
                                # @TODO code to change Frequency
                                pass
                            else:
                                messagebox.showwarning('Frequency', f'Set Frequency to {f}kHz')
                                if not firstFreq:
                                    cooldownTime = 300
                                    while cooldownTime > 0:
                                        sleep(1)
                                        cooldownTime -= 1
                                        ctmin = int(cooldownTime // 60)
                                        ctsec = int(cooldownTime % 60)
                                        self.statusLabel['text'] = f'Cooling down {ctmin:2d}:{ctsec:02d}'
                                firstFreq = 0

                            dataVin = list()
                            for vin in self.pvinList:
                                # Changing Vin
                                if eval(ConfigMgr.instr['vinPSOnOff']):
                                    if self.vin is not None:
                                        self.statusLabel['text'] = f'Setting Vin to {vin}kHz'
                                        self.vin.ch.Configure_VoltageLevel(Level=vin, CurrentLimit=20)
                                        self.vin.ch.Enable(True)
                                else:
                                    messagebox.showwarning('Vin', f'Set Vin to {vin}V')

                                dataList = list()
                                deadRise = 0
                                deadFall = 0
                                for i in self.currList:
                                    mdate = dt.datetime.now()
                                    mydate = mdate.strftime('%m/%d/%Y')
                                    mytime = mdate.strftime('%H-%M-%S')
                                    if self.scope is not None:
                                        self.scope.run()  # Run
                                        self.scope.Trigger_Edge(
                                            Level=2.00,
                                            Slope='RISE',
                                            Position=2,
                                            Coupling='DC',
                                            ChannelIndex=int(swCh)
                                        )

                                    # Set the current
                                    if self.load is not None:
                                        self.statusLabel['text'] = f'Setting Load to {i}A'
                                        self.load.set_channel(ConfigMgr.instr['loadChannel'][-1])
                                        self.load.set_mode('CCH')
                                        self.load.set_current('STATIC', i)
                                        if i == 0:
                                            self.load.output('OFF')
                                        else:
                                            self.load.output('ON')
                                    startTime = dt.datetime.now()
                                    if (self.scope is not None) and (i >= self.currList[-1] / 2):
                                        self.scope.reset_statistics()

                                        self.scope.set_vertical_scale(ch=swCh, value=0.3)
                                        self.scope.verpos(swCh, 0)
                                        self.scope.set_record_length(1000000)
                                        self.scope.horpos(75)
                                        sleep(1.0)

                                        # Stop Scope
                                        self.scope.stop()
                                        sleep(0.5)

                                        pointA = 0
                                        pointB = 0
                                        wave = self.scopeSwCh.GetProbeWaveform(Timeout=0.5)
                                        if self.scope.ID == '4k':
                                            slicedWave = wave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                            pointA = slicedWave.Measurements_Utils.argmin(return_index=True)
                                            crossZeroA = aux.find_rise_dead_time(slicedWave, pointA,
                                                                                direction='BACKWARD')
                                            crossZeroB = aux.find_rise_dead_time(slicedWave, pointA,
                                                                                direction='FORWARD')
                                            if crossZeroA is None:
                                                print('PointA for dead time rise not found')
                                                crossZeroA = (0, 0, 0)
                                            if crossZeroB is None:
                                                print('PointB for dead time rise not found')
                                                crossZeroB = (0, 0, 0)
                                            pointA = crossZeroA[2]
                                            pointB = crossZeroB[2]
                                        else:
                                            slicedWave = wave
                                            pointA = slicedWave.Measurements_Utils.argmin(return_index=True)
                                            crossZeroA = aux.find_rise_dead_time(slicedWave, pointA,
                                                                                direction='BACKWARD')
                                            crossZeroB = aux.find_rise_dead_time(slicedWave, pointA,
                                                                                direction='FORWARD')
                                            if crossZeroA is None:
                                                print('PointA for dead time rise not found')
                                                crossZeroA = (0, 0, 0)
                                            if crossZeroB is None:
                                                print('PointB for dead time rise not found')
                                                crossZeroB = (0, 0, 0)
                                            pointA = crossZeroA[2] - (slicedWave.index_to_time(-1) * 0.75)
                                            pointB = crossZeroB[2] - (slicedWave.index_to_time(-1) * 0.75)

                                        if pointA == None:
                                            pointA = 0
                                        if pointB == None:
                                            pointB = 0

                                        self.scope.set_cursor_pos(cursor=1, pos=pointA)
                                        self.scope.set_cursor_pos(cursor=2, pos=pointB)
                                        sleep(0.5)
                                        deadRise = pointB - pointA

                                        # Waveform
                                        fileName_i = folderName + f'/images/deadTime_rise_{t}C_{vcc}Vcc_{vout}Vin_{vin}Vout_{f}kHz_{i}A.png'
                                        self.scope.saveimage(fileName_i, Background='white')

                                        self.scope.run()
                                        sleep(0.1)
                                        self.scope.horpos(25)

                                        self.scope.Trigger_Edge(
                                            Level=2.0,
                                            Slope='FALL',
                                            Position=2,
                                            Coupling='DC',
                                            ChannelIndex=int(swCh)
                                        )
                                        self.scope.reset_statistics()
                                        sleep(1.0)

                                        self.scope.stop()
                                        sleep(0.5)
                                        wave = self.scopeSwCh.GetProbeWaveform(Timeout=0.5)
                                        if self.scope.ID == '4k':
                                            slicedWave = wave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
                                            slicedWave2 = wave.slice_by_time(start_time=6e-9, stop_time=12e-9)
                                            pointA = slicedWave.Measurements_Utils.find_nth_crossing(0, "fall", 0)
                                            pointB = slicedWave2.Measurements_Utils.argmax()
                                        else:
                                            slicedWave = wave
                                            slicedWave2 = wave.slice_by_time(start_time=wave.index_to_time(-1) * 0.25 + 6e-9, stop_time=wave.index_to_time(-1) * 0.25 + 12e-9)
                                            pointA = slicedWave.Measurements_Utils.find_nth_crossing(0, "fall", 0) - (slicedWave.index_to_time(-1) * 0.25)
                                            pointB = slicedWave2.Measurements_Utils.find_nth_crossing(0, 'rise', 0) - (slicedWave.index_to_time(-1) * 0.25)
                                            if pointB is None:
                                                pointB = slicedWave2.Measurements_Utils.argmax() - (slicedWave.index_to_time(-1) * 0.25)

                                        if pointA == None:
                                            pointA = 0
                                        if pointB == None:
                                            pointB = 0

                                        self.scope.set_cursor_pos(cursor=1, pos=pointA)
                                        self.scope.set_cursor_pos(cursor=2, pos=pointB)
                                        sleep(0.5)
                                        deadFall = pointB - pointA

                                        # Waveform
                                        fileName_i = folderName + f'/images/deadTime_fall_{t}C_{vcc}Vcc_{vout}Vin_{vin}Vout_{f}kHz_{i}A.png'
                                        self.scope.saveimage(fileName_i, Background='white')

                                    endTime = dt.datetime.now()

                                    img_delay_time = endTime - startTime
                                    if img_delay_time.total_seconds() > eval(ConfigMgr.testConditions101['soak']):
                                        pass
                                    else:
                                        sleep(eval(ConfigMgr.testConditions101['soak']) - img_delay_time.total_seconds())

                                    # print('processing time =', img_delay_time.total_seconds(), 'soaking time left =', wait_time - img_delay_time.total_seconds())

                                    if self.keith is not None:
                                        results = self.keith.read(num=20)
                                        # Parsing
                                        str_lst = results.split(',')
                                        float_lst = [float(x) for x in str_lst]
                                    try:
                                        kiout = float_lst[eval(ConfigMgr.testConditions101['kiout']) - 1]
                                    except:
                                        kiout = 0
                                    try:
                                        iload = 1000 * kiout / eval(ConfigMgr.testConditions101['kioutr'])
                                    except:
                                        iload = 0
                                    try:
                                        kvin = float_lst[eval(ConfigMgr.testConditions101['kvin']) - 1]
                                    except:
                                        kvin = 0
                                    try:
                                        tmon = float_lst[eval(ConfigMgr.testConditions101['ktmon']) - 1]
                                    except:
                                        tmon = 0
                                    try:
                                        kvout = float_lst[eval(ConfigMgr.testConditions101['kvout']) - 1]
                                    except:
                                        kvout = 0
                                    try:
                                        imon = float_lst[eval(ConfigMgr.testConditions101['kimon']) - 1] * 1000
                                    except:
                                        imon = 0
                                    try:
                                        siin = float_lst[eval(ConfigMgr.testConditions101['kiin']) - 1]
                                    except:
                                        siin = 0
                                    try:
                                        iiin = 1000 * siin / eval(ConfigMgr.testConditions101['kiinr'])
                                    except:
                                        iiin = 0
                                    try:
                                        kvcc = float_lst[eval(ConfigMgr.testConditions101['kvcc']) - 1]
                                    except:
                                        kvcc = 0
                                    try:
                                        sicc = float_lst[eval(ConfigMgr.testConditions101['kicc']) - 1]
                                    except:
                                        sicc = 0
                                    try:
                                        iicc = 1000 * sicc / eval(ConfigMgr.testConditions101['kiccr'])
                                    except:
                                        iicc = 0
                                    try:
                                        kpgood = float_lst[eval(ConfigMgr.testConditions101['kpgook']) - 1]
                                    except:
                                        kpgood = 0

                                    pvcc = kvcc * iicc
                                    pin = kvin * iiin
                                    pout = kvout * iload
                                    try:
                                        effeciency = (pout / (pvcc + pin)) * 100
                                    except:
                                        effeciency = 0
                                    iout = iload
                                    temp = 0

                                    # Table update info
                                    if tableID % 2 == 0:
                                        self.table.insert(
                                            parent='',
                                            index='end',
                                            iid=tableID,
                                            text='',
                                            values=(mydate, mytime, kvin, iiin, kvcc, iicc, kvout, kiout, effeciency),
                                            tags=('evenrow',)
                                        )
                                    else:
                                        self.table.insert(
                                            parent='',
                                            index='end',
                                            iid=tableID,
                                            text='',
                                            values=(mydate, mytime, kvin, iiin, kvcc, iicc, kvout, kiout, effeciency),
                                            tags=('oddrow',)
                                        )
                                    tableID += 1

                                    self.yAxis.append(float(effeciency))
                                    self.xAxis.append(float(i))
                                    self.a.plot(self.xAxis, self.yAxis)
                                    self.plotCanvas.draw_idle()

                                    dataList.append(
                                        {
                                            f'{i}A': {
                                                'vcc_step': vcc,
                                                'vin_step': vin,
                                                'vout_step': vout,
                                                'freq_step': f,
                                                'kiout': kiout,
                                                'kvout': kvout,
                                                'iload': iload,
                                                'kvin': kvin,
                                                'siin': siin,
                                                'iiin': iiin,
                                                'kvcc': kvcc,
                                                'sicc': sicc,
                                                'iicc': iicc,
                                                'tmon': tmon,
                                                'kpgood': kpgood,
                                                'pin': pin,
                                                'pvcc': pvcc,
                                                'pout': pout,
                                                'effeciency': effeciency,
                                                'temp': temp,
                                                'imon': imon,
                                                'deadRise': deadRise,
                                                'deadFall': deadFall,
                                            }
                                        }
                                    )

                                    # Store the data in the Excel sheet (in memory until sheet is saved)...
                                    ws1.cell(column=4, row=RowIndex, value=vcc)
                                    ws1.cell(column=5, row=RowIndex, value=mydate)
                                    ws1.cell(column=6, row=RowIndex, value=mytime)
                                    ws1.cell(column=7, row=RowIndex, value=vin)
                                    ws1.cell(column=8, row=RowIndex, value=vout)
                                    ws1.cell(column=9, row=RowIndex, value=f)
                                    ws1.cell(column=10, row=RowIndex, value=i)

                                    ws1.cell(column=12, row=RowIndex, value=kvin)
                                    ws1.cell(column=13, row=RowIndex, value=(siin * 1000))
                                    ws1.cell(column=14, row=RowIndex, value=iiin)
                                    ws1.cell(column=15, row=RowIndex, value=kvcc)
                                    ws1.cell(column=16, row=RowIndex, value=(sicc * 1000))
                                    ws1.cell(column=17, row=RowIndex, value=iicc)
                                    ws1.cell(column=18, row=RowIndex, value=kvout)
                                    ws1.cell(column=19, row=RowIndex, value=(kiout * 1000))
                                    ws1.cell(column=20, row=RowIndex, value=iload)
                                    ws1.cell(column=21, row=RowIndex, value=tmon)
                                    ws1.cell(column=22, row=RowIndex, value=kpgood)
                                    ws1.cell(column=23, row=RowIndex, value=pin)
                                    ws1.cell(column=24, row=RowIndex, value=pvcc)
                                    ws1.cell(column=25, row=RowIndex, value=pout)
                                    ws1.cell(column=26, row=RowIndex, value=(pin + pvcc) - pout)
                                    ws1.cell(column=27, row=RowIndex, value=effeciency)
                                    ws1.cell(column=28, row=RowIndex, value=temp)
                                    ws1.cell(column=29, row=RowIndex, value=imon)
                                    ws1.cell(column=30, row=RowIndex, value=(deadRise * 1e+9))
                                    ws1.cell(column=31, row=RowIndex, value=(deadFall * 1e+9))

                                    # Advance to the next data row...
                                    RowIndex = RowIndex + 1
                                    fileName = folderName + '/data/efficiency.xlsx'
                                    wb.save(filename=fileName)
                                # Scope
                                if self.scope is not None:
                                    self.scope.run()
                                if self.load is not None:
                                    self.load.set_current('STATIC', 0)

                                dataVin.append({f'{vin}Vin': dataList})
                            if self.vin is not None:
                                self.vin.ch.Enable(False)
                            RowIndex += 1
                            dataF.append({f'{f}kHz': dataVin})
                        dataVcc.append({f'{vcc}Vcc': dataF})
                    dataVout.append({f'{vout}Vout': dataVcc})
                test101Dict[f'{t}C'] = dataVout

                # Save the output file...
                fileName = folderName + '/data/efficiency.xlsx'
                wb.save(filename=fileName)

                with open(folderName + '/data/efficiency_test.json', 'w') as f:
                    json.dump(test101Dict, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(e)
            if self.load is not None:
                self.load.set_current('STATIC', 0)
                self.load.output('OFF')

        # region SETUP LIST VARIABLES

    def setting_pre_test_variables(self, config):
        self.tempList = list()
        self.currList = list()
        self.freqList = list()
        self.pvinList = list()
        self.voutList = list()
        self.vccList = list()
        # Temperature
        for key in ConfigMgr.tempSteps:
            if ConfigMgr.tempSteps[key] == '':
                continue
            self.tempList.append(eval(ConfigMgr.tempSteps[key]))
        # Current / Load
        startCurrent = eval(config['startCurr'])
        endCurrent = eval(config['endCurr'])
        stepCurrent = eval(config['stepCurr'])
        startCurrent *= 10
        endCurrent *= 10
        stepCurrent *= 10
        for _ in range(int(startCurrent), int(endCurrent) + int(stepCurrent), int(stepCurrent)):
            if _ > int(endCurrent):
                _ = int(endCurrent)
            self.currList.append(_ / 10.0)
        # Frequency
        startFreq = eval(config['startFsw'])
        endFreq = eval(config['endFsw'])
        try:
            stepFreq = eval(config['stepFsw'])
        except:
            stepFreq = 0
        if stepFreq == 0:
            self.freqList = [startFreq]
        else:
            for _ in range(int(startFreq), int(endFreq) + int(stepFreq), int(stepFreq)):
                if _ > int(endFreq):
                    _ = int(endFreq)
                self.freqList.append(_)
        # Vin
        startPVin = eval(config['startPVin'])
        endPVin = eval(config['endPVin'])
        try:
            stepPVin = eval(config['stepPVin'])
        except:
            stepPVin = 0
        if stepPVin == 0:
            self.pvinList = [startPVin]
        else:
            startPVin *= 10
            endPVin *= 10
            stepPVin *= 10
            for _ in range(int(startPVin), int(endPVin) + int(stepPVin), int(stepPVin)):
                if _ > int(endPVin):
                    _ = int(endPVin)
                self.pvinList.append(_ / 10.0)
        # Vout
        startVout = eval(config['startVout'])
        endVout = eval(config['endVout'])
        try:
            stepVout = eval(config['stepVout'])
        except:
            stepVout = 0
        startVout *= 10
        endVout *= 10
        stepVout *= 10
        if stepVout == 0:
            self.voutList = [eval(config['startVout'])]
        else:
            for i in range(int(startVout), int(endVout) + int(stepVout), int(stepVout)):
                if i > int(endVout):
                    i = int(endVout)
                self.voutList.append(i / 10.0)
        # Vcc
        startVcc = eval(config['startVcc'])
        endVcc = eval(config['endVcc'])
        try:
            stepVcc = eval(config['stepVcc'])
        except:
            stepVcc = 0
        startVcc *= 10
        endVcc *= 10
        stepVcc *= 10
        if stepVcc == 0:
            self.vccList = [eval(config['startVcc'])]
        else:
            for v in range(int(startVcc), int(endVcc) + int(stepVcc), int(stepVcc)):
                if v > int(endVcc):
                    v = int(endVcc)
                self.vccList.append(v / 10.0)

    # endregion

    # region MAKE FOLDER
    def make_dir(self, name: str):
        date = dt.datetime.now()
        date = date.strftime("%Y_%m_%d_%H_%M_%S")
        test = name + date
        isDataDir = os.path.isdir('polData')
        if (not isDataDir):
            os.mkdir('polData')
        isTestDir = os.path.isdir('polData/' + str(test))
        if (not isTestDir):
            os.mkdir('polData/' + str(test))
        isTestImgDir = os.path.isdir('polData/' + str(test) + '/images')
        if (not isTestImgDir):
            os.mkdir('polData/' + str(test) + '/images')
        isTestDataDir = os.path.isdir('polData/' + str(test) + '/data')
        if (not isTestDataDir):
            os.mkdir('polData/' + str(test) + '/data')
        return 'polData/' + test
    # endregion

    def create_excel(self, title='', index=0, config=None):
        # Create Excel File
        wb = Workbook()
        ws1 = wb.create_sheet(title=title, index=index)

        # Set up Excel Doc:
        # Parameter ...

        ws1.column_dimensions['A'].width = 24
        ws1.column_dimensions['B'].width = 9
        r = 2
        ws1.cell(column=1, row=r, value="Start Current (A):")
        ws1.cell(column=2, row=r, value=config['startCurr'])
        r += 1
        ws1.cell(column=1, row=r, value="Final Current(A):")
        ws1.cell(column=2, row=r, value=config['endCurr'])
        r += 1
        ws1.cell(column=1, row=r, value="Increments (A):")
        ws1.cell(column=2, row=r, value=config['stepCurr'])
        r += 1
        ws1.cell(column=1, row=r, value="Time for statistics count (s):")
        ws1.cell(column=2, row=r, value=eval(config['soak']) * 60)
        r += 1
        ws1.cell(column=1, row=r, value="Vdrv list (V):")
        ws1.cell(column=2, row=r, value=str(self.vccList))
        r += 1
        ws1.cell(column=1, row=r, value="Vin list (V):")
        ws1.cell(column=2, row=r, value=str(self.pvinList))
        r += 1
        ws1.cell(column=1, row=r, value="Fsw list (kHz):")
        ws1.cell(column=2, row=r, value=str(self.freqList))
        r += 1
        ws1.cell(column=1, row=r, value="Vout list (V):")
        ws1.cell(column=2, row=r, value=str(self.voutList))
        # data
        if title == 'Vout Ripple Fsw vs Load':
            self.data104_excel(ws1)
        if title == 'Efficiency':
            self.data101_excel(ws1)
        return wb

    def data101_excel(self, ws1):
        c = 4
        ws1.cell(column=c, row=2, value='VCC')
        c += 1
        ws1.cell(column=c, row=2, value='Date')
        c += 1
        ws1.cell(column=c, row=2, value='Time')
        c += 1
        ws1.cell(column=c, row=2, value='Pvin (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Vout (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Frequency (kHz)')
        c += 1
        ws1.cell(column=c, row=2, value='Load (A)')
        c += 2
        ws1.cell(column=c, row=2, value='Vin (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Iin (mV)')
        c += 1
        ws1.cell(column=c, row=2, value='Iin (A)')
        c += 1
        ws1.cell(column=c, row=2, value='Vcc (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Icc (mV)')
        c += 1
        ws1.cell(column=c, row=2, value='Icc (A)')
        c += 1
        ws1.cell(column=c, row=2, value='Vout (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Iout (mV)')
        c += 1
        ws1.cell(column=c, row=2, value='Iout (A)')
        c += 1
        ws1.cell(column=c, row=2, value='Tmon (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Pgood (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Pin (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Pvcc (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Pout (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Power Loss (W)')
        c += 1
        ws1.cell(column=c, row=2, value='Efficiency (%)')
        c += 1
        ws1.cell(column=c, row=2, value='Tmon (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Imon (V)')
        c += 1
        ws1.cell(column=c, row=2, value='Dead Time Rise (ns)')
        c += 1
        ws1.cell(column=c, row=2, value='Dead Time Fall (ns)')
