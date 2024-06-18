import matplotlib.pyplot as plt
import datetime as dt
import os
from openpyxl.styles import Alignment


def find_crossing_zero(wave, startPoint, endPoint=None, direction='forward', *args, **kwargs):
    '''
    :param wave: Waveform produced by Ifx-PyVerify libraries
    :param startPoint: index of starting point
    :param direction: { FORWARD | BACKWARD }
    :return: tuple (index, data, time) | None if no crossing found
    '''
    if direction.upper() == 'BACKWARD':
        step = -1
        if endPoint is None:
            endPoint = 0
    else:
        step = 1
        if endPoint is None:
            endPoint = len(wave.data)

    if wave.data[startPoint] < 0:
        for idx in range(startPoint, endPoint, step):
            if wave.data[idx] >= 0:
                return (idx, wave.data[idx], wave.time[idx])
    elif wave.data[startPoint] > 0:
        for idx in range(startPoint, endPoint, step):
            if wave.data[idx] <= 0:
                return (idx, wave.data[idx], wave.time[idx])
    else:
        idx2 = startPoint + step
        if wave.data[idx2] > 0:
            for idx in range(startPoint, endPoint, step):
                if wave.data[idx] <= 0:
                    return (idx, wave.data[idx], wave.time[idx])
        elif wave.data[idx2] < 0:
            for idx in range(startPoint, endPoint, step):
                if wave.data[idx] >= 0:
                    return (idx, wave.data[idx], wave.time[idx])
    return None

def find_rise_dead_time(wave, startPoint, endPoint=None, direction='forward', *args, **kwargs):
    '''
    :param wave: Waveform produced by Ifx-PyVerify libraries
    :param startPoint: index of starting point
    :param direction: { FORWARD | BACKWARD }
    :return: tuple (index, data, time) | corner if no crossing zero found
    '''
    if direction.upper() == 'BACKWARD':
        step = -1
        if endPoint is None:
            endPoint = 0
    else:
        step = 1
        if endPoint is None:
            endPoint = len(wave.data)

    if wave.data[startPoint] < 0:
        first = 0
        corner = (0,0,0)
        currMeas = wave.data[startPoint]
        prevMeas = currMeas
        for idx in range(startPoint, endPoint, step):
            currMeas = wave.data[idx]
            if wave.data[idx] >= 0:
                return (idx, wave.data[idx], wave.time[idx])
            if currMeas - prevMeas < 0 and not first:
                first = 1
                corner = (idx, wave.data[idx-step], wave.time[idx-step])
            prevMeas = currMeas
    elif wave.data[startPoint] > 0:
        first = 0
        corner = (0, 0, 0)
        currMeas = wave.data[startPoint]
        prevMeas = currMeas
        for idx in range(startPoint, endPoint, step):
            currMeas = wave.data[idx]
            if wave.data[idx] <= 0:
                return (idx, wave.data[idx], wave.time[idx])
            if currMeas - prevMeas < 0 and not first:
                first = 1
                corner = (idx, wave.data[idx-step], wave.time[idx-step])
            prevMeas = currMeas
    else:
        idx2 = startPoint + step
        if wave.data[idx2] > 0:
            first = 0
            corner = (0, 0, 0)
            currMeas = wave.data[startPoint]
            prevMeas = currMeas
            for idx in range(startPoint, endPoint, step):
                currMeas = wave.data[idx]
                if wave.data[idx] <= 0:
                    return (idx, wave.data[idx], wave.time[idx])
                if currMeas - prevMeas < 0 and not first:
                    first = 1
                    corner = (idx, wave.data[idx - step], wave.time[idx - step])
                prevMeas = currMeas
        elif wave.data[idx2] < 0:
            first = 0
            corner = (0, 0, 0)
            currMeas = wave.data[startPoint]
            prevMeas = currMeas
            for idx in range(startPoint, endPoint, step):
                currMeas = wave.data[idx]
                if wave.data[idx] >= 0:
                    return (idx, wave.data[idx], wave.time[idx])
                if currMeas - prevMeas < 0 and not first:
                    first = 1
                    corner = (idx, wave.data[idx - step], wave.time[idx - step])
                prevMeas = currMeas
    return corner

def setting_pre_test_variables(config, configT, tempList, currList, currList2, freqList, pvinList, voutList, vccList, modeList, enList, biasList, delayList, *args, **kwargs):
    tempList.clear()
    currList.clear()
    currList2.clear()
    freqList.clear()
    pvinList.clear()
    voutList.clear()
    vccList.clear()
    modeList.clear()
    enList.clear()
    biasList.clear()
    delayList.clear()
    # Temperature
    for key in configT:
        if configT[key] == '':
            continue
        tempList.append(eval(configT[key]))
    # Current / Load
    if config['currOpt'] == 'fixed':
        for key in ['curr1', 'curr2', 'curr3', 'curr4', 'curr5', 'curr6']:
            if config[key] != '':
                currList.append(eval(config[key]))
    else:
        startCurrent = eval(config['startCurr'])
        try:
            stepCurrent = eval(config['stepCurr'])
            endCurrent = eval(config['endCurr'])
        except:
            stepCurrent = 0
            endCurrent = 0

        if stepCurrent == 0:
            currList.append(startCurrent/10.0)
        else:
            startCurrent *= 10
            endCurrent *= 10
            stepCurrent *= 10
            for _ in range(int(startCurrent), int(endCurrent) + int(stepCurrent), int(stepCurrent)):
                if _ > int(endCurrent):
                    _ = int(endCurrent)
                currList.append(_ / 10.0)
    # Current2 / Load2
    startCurrent = eval(config['startCurr2'])
    try:
        stepCurrent = eval(config['stepCurr2'])
        endCurrent = eval(config['endCurr2'])
    except:
        stepCurrent = 0
        endCurrent = 0

    if stepCurrent == 0:
        currList2.append(startCurrent / 10.0)
    else:
        startCurrent *= 10
        endCurrent *= 10
        stepCurrent *= 10
        for _ in range(int(startCurrent), int(endCurrent) + int(stepCurrent), int(stepCurrent)):
            if _ > int(endCurrent):
                _ = int(endCurrent)
            currList2.append(_ / 10.0)
    # Frequency
    if config['fswOpt'] == 'fixed':
        for key in ['fsw1', 'fsw2', 'fsw3', 'fsw4', 'fsw5', 'fsw6', 'fsw7', 'fsw8', 'fsw9']:
            if config[key] != '':
                freqList.append(eval(config[key]))
    else:
        startFreq = eval(config['startFsw'])
        try:
            stepFreq = eval(config['stepFsw'])
            endFreq = eval(config['endFsw'])
        except:
            stepFreq = 0
            endFreq = 0
        if stepFreq == 0:
            freqList.append(startFreq)
        else:
            for _ in range(int(startFreq), int(endFreq) + int(stepFreq), int(stepFreq)):
                if _ > int(endFreq):
                    _ = int(endFreq)
                freqList.append(_)
    # Vin
    if config['pvinOpt'] == 'fixed':
        for key in ['PVin1', 'PVin2', 'PVin3', 'PVin4', 'PVin5', 'PVin6']:
            try:
                if config[key] == '':
                    continue
                else:
                    pvinList.append(config[key])
            except:
                continue
    else:
        startPVin = eval(config['startPVin'])
        try:
            stepPVin = eval(config['stepPVin'])
            endPVin = eval(config['endPVin'])
        except:
            stepPVin = 0
            endPVin = 0
        if stepPVin == 0:
            pvinList.append(startPVin)
        else:
            startPVin *= 100
            endPVin *= 100
            stepPVin *= 100
            for _ in range(int(startPVin), int(endPVin) + int(stepPVin), int(stepPVin)):
                if _ > int(endPVin):
                    _ = int(endPVin)
                pvinList.append(_ / 100.0)
    # Vout
    if config['voutOpt'] == 'fixed':
        for key in ['vout1', 'vout2', 'vout3', 'vout4', 'vout5', 'vout6']:
            try:
                if config[key] == '':
                    continue
                else:
                    voutList.append(config[key])
            except:
                continue
    else:
        startVout = eval(config['startVout'])
        try:
            stepVout = eval(config['stepVout'])
            endVout = eval(config['endVout'])
        except:
            stepVout = 0
            endVout = 0
        if stepVout == 0:
            voutList.append(eval(config['startVout']))
        else:
            startVout *= 10
            endVout *= 10
            stepVout *= 10
            for i in range(int(startVout), int(endVout) + int(stepVout), int(stepVout)):
                if i > int(endVout):
                    i = int(endVout)
                voutList.append(i / 10.0)
    # Vcc
    if eval(config['vccLdo']):
        vccList.append(-1)
    if eval(config['vccExt']):
        startVcc = eval(config['startVcc'])
        try:
            stepVcc = eval(config['stepVcc'])
            endVcc = eval(config['endVcc'])
        except:
            stepVcc = 0
            endVcc = 0

        if stepVcc == 0:
            vccList.append(startVcc)
        else:
            startVcc *= 100
            endVcc *= 100
            stepVcc *= 100
            for v in range(int(startVcc), int(endVcc) + int(stepVcc), int(stepVcc)):
                if v > int(endVcc):
                    v = int(endVcc)
                vccList.append(v / 100.0)
    # enList
    if config['enOpt'] == 'EN':
        enList.append(-1)
    else:
        for key in ['en1', 'en2', 'en3']:
            try:
                if config[key] == '':
                    continue
                enList.append(config[key])
            except:
                continue
    # biasList
    if config['biasOpt'] == '0':
        biasList.append(-1)
    else:
        for key in ['bias1', 'bias2']:
            try:
                if config[key] == '':
                    continue
                biasList.append(config[key])
            except:
                continue
    # delayList
    for key in ['delay1', 'delay2', 'delay3']:
        try:
            if config[key] == '' or config[key] == '0':
                continue
            else:
                delayList.append(eval(config[key]) * 1e-3)
        except:
            continue
    # MODE DEM and FCCM
    if eval(config['modeDEM']):
        modeList.append('DEM')
    if eval(config['modeFCCM']):
        modeList.append('FCCM')

def make_dir(name: str):
    date = dt.datetime.now()
    date = date.strftime("%Y_%m_%d_%H_%M_%S")
    test = date + '_' + name
    isDataDir = os.path.isdir('Data')
    if (not isDataDir):
        os.mkdir('Data')
    isTestDir = os.path.isdir('Data/' + str(test))
    if (not isTestDir):
        os.mkdir('Data/' + str(test))
    isTestImgDir = os.path.isdir('Data/' + str(test) + '/images')
    if (not isTestImgDir):
        os.mkdir('Data/' + str(test) + '/images')
    isTestDataDir = os.path.isdir('Data/' + str(test) + '/data')
    if (not isTestDataDir):
        os.mkdir('Data/' + str(test) + '/data')
    return 'Data/' + test

def excel_info(ws1, config, currList, vccList, pvinList, freqList, voutList, bomtxt):
    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 26
    r = 2
    ws1.cell(column=1, row=r, value=f'Load Steps:')
    ws1.cell(column=2, row=r, value=str(currList))
    r += 1
    ws1.cell(column=1, row=r, value="Time for statistics count (s):")
    ws1.cell(column=2, row=r, value=eval(config['soak']) * 60)
    r += 1
    if config['vccLdo'] == '1':
        ws1.cell(column=1, row=r, value="VCC list (V):")
        ws1.cell(column=2, row=r, value='LDO')
        r += 1
    if config['vccExt'] == '1':
        ws1.cell(column=1, row=r, value="VCC list (V):")
        ws1.cell(column=2, row=r, value=str(vccList))
        r += 1
    ws1.cell(column=1, row=r, value="Vin list (V):")
    ws1.cell(column=2, row=r, value=str(pvinList))
    r += 1
    ws1.cell(column=1, row=r, value="Fsw list (kHz):")
    ws1.cell(column=2, row=r, value=str(freqList))
    r += 1
    ws1.cell(column=1, row=r, value="Vout list (V):")
    ws1.cell(column=2, row=r, value=str(voutList))
    r += 1
    ws1.cell(column=1, row=r, value="Test Conditions:")
    r += 1
    ws1.merge_cells(start_row=r, end_row=r + 15, start_column=1, end_column=1)
    ws1.cell(column=1, row=r).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws1.cell(column=1, row=r, value=bomtxt)


if __name__ == '__main__':
    scope = Tkdpo4k('USB0::0x0699::0x0401::C000576::INSTR', Reset=False, Simulate=False)
    scope.swCh = scope.GetChannel(Index=4)
    scope.horizontal_delay(state='OFF')
    wave = scope.swCh.GetProbeWaveform(Timeout=0.5)
    slicedWave = wave.slice_by_time(start_time=-20e-9, stop_time=20e-9)
    pointA = slicedWave.Measurements_Utils.argmin(return_index=True)
    print(pointA)
    crossZeroA = find_rise_dead_time(slicedWave, pointA, direction='BACKWARD')
    crossZeroB = find_rise_dead_time(slicedWave, pointA, direction='FORWARD')
    if crossZeroA is None:
        crossZeroA = (0,)
    plt.plot(slicedWave.time, slicedWave.data)
    plt.axvline(slicedWave.time[crossZeroA[0]], color='r')
    plt.axvline(slicedWave.time[crossZeroB[0]], color='r')
    plt.show()

