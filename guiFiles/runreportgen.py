"""@package docstring
This is a report generating file for standardized POL/power stage data files;
Standardization based upon POL GUI user guide.

User must select files to be generated to a report when prompted.
Report will contained organized tables based upon data available, and graphical representations of each table.
User may alter graphs for data analysis after report generation is complete.

Version number: 0.01.00
Last updated: 08/11/2022
"""

# Imports
import tkinter as tk
from tkinter import filedialog
import os
import openpyxl as xl
import openpyxl.utils.cell as util
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.styles import (
    Alignment,
    PatternFill,
    Border,
    Side
)

# Global Constant
# The maximum column number to map.
MAXCOLUMN = 20


# Class containing all relevant function for POL/power stage standardized report generation.
# Inputs: None.
class RunReportGen:


    # Functions
    def __init__(self):
        """! Initializes the program: initialize lists as lists and object variables as None.

        Inputs: None.
        """
        self.destFolder = '.'
        self.finalReport = 'FinalReport.xlsx'
        self.col = None
        self.descList = list()
        self.row = None
        self.currentRow = None
        self.wb = None
        self.root = None
        self.filePaths = None
        self.currentCol = None
        self.tableList = list()

        self.testSpecificFunc = {
            '101' : self.report101,
            '104' : self.report104,
            '201' : self.report201,
            '202' : self.report202,
            '203' : self.report203,
            '307' : self.report307,
            'eff' : self.efficiency,
            'sw' : self.switch,
        }

    def _get_file_paths(self):
        """! Opens window to located filepath.
        Inputs: None.
        """

        self.root = tk.Tk()
        self.root.withdraw()  # Remove tkinter window.

        # Selection of filepaths only in csv format.
        self.filePaths = filedialog.askopenfilenames(
            defaultextension='.csv',
            filetypes=(('csv files', '*.csv'),),
        )

    def report101(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0].replace('@&', '\n')
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom

            numOfCols = int(valList[3])
            numOfRows = int(valList[2])

            return numOfRows, numOfCols

    def report104(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0]
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom.replace('@&', '\n')

            ws.cell(row=1, column=3).value = f'DEM Boundary'
            ws.cell(row=2, column=3).value = valList[1]
            ws.cell(row=1, column=3).border = Border(top=med, bottom=med, left=med, right=med)
            ws.cell(row=2, column=3).border = Border(top=med, bottom=med, left=med, right=med)

            numOfCols = int(valList[3])
            numOfRows = int(valList[2])

            return numOfRows, numOfCols

    def report201(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0].replace('@&', '\n')
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom

            numOfCols = int(valList[3])
            numOfRows = int(valList[2])

            return numOfRows, numOfCols

    def report202(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0].replace('@&', '\n')
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom

            numOfCols = int(valList[3])
            numOfRows = int(valList[2])

            return numOfRows, numOfCols

    def report203(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0].replace('@&', '\n')
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom

            numOfCols = int(valList[3])
            numOfRows = int(valList[2])

            return numOfRows, numOfCols

    def report307(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0]
            resVal = valList[1]
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom.replace('@&', '\n')

            ws.cell(row=1, column=3).value = f'CS Resistor'
            ws.cell(row=2, column=3).value = resVal
            ws.cell(row=1, column=3).border = Border(top=med, bottom=med, left=med, right=med)
            ws.cell(row=2, column=3).border = Border(top=med, bottom=med, left=med, right=med)

            ws.cell(row=1, column=4).value = f'CS Gain Min'
            ws.cell(row=2, column=4).value = valList[4]
            ws.cell(row=1, column=4).border = Border(top=med, bottom=med, left=med, right=med)
            ws.cell(row=2, column=4).border = Border(top=med, bottom=med, left=med, right=med)

            ws.cell(row=1, column=5).value = f'CS Gain Typ'
            ws.cell(row=2, column=5).value = valList[5]
            ws.cell(row=1, column=5).border = Border(top=med, bottom=med, left=med, right=med)
            ws.cell(row=2, column=5).border = Border(top=med, bottom=med, left=med, right=med)

            ws.cell(row=1, column=6).value = f'CS Gain Max'
            ws.cell(row=2, column=6).value = valList[6]
            ws.cell(row=1, column=6).border = Border(top=med, bottom=med, left=med, right=med)
            ws.cell(row=2, column=6).border = Border(top=med, bottom=med, left=med, right=med)
            numOfCols = int(valList[3])
            numOfRows = int(valList[2])
            return numOfRows, numOfCols

    def efficiency(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0]
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom.replace('@&', '\n')

            numOfCols = int(valList[3])
            numOfRows = int(valList[2])
            return numOfRows, numOfCols

    def switch(self, line, med, valList, sheetName, numOfCols):
        if line:
            for i in range(3, numOfCols + 4):
                tableName = valList[i]
                self.tableList.append(tableName)
        else:
            bom = valList[0]
            ws = self.wb.create_sheet(title=f'{sheetName}_Comments')
            ws.column_dimensions['A'].width = 32
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=2, column=1).value = bom.replace('@&', '\n')

            numOfCols = int(valList[3])
            numOfRows = int(valList[2])
            return numOfRows, numOfCols

    def generic_report_gen(self, test=''):
        """! Main program entry:
            Selects files.
            Writes csv files to excel workbook.
            Plot graphs according to data.

            Inputs: None.
            """

        self.wb = xl.Workbook()  # Open a workbook to write and save csv data to.
        med = Side(border_style="medium")

        # Opening a singular file in the list of desired files.
        # Inputs: None.
        for path in self.filePaths:
            firstRun = 0
            self.tableList.clear()
            self.descList.clear()

            # Opening file in reading mode.
            # Inputs: None.
            with open(path, 'r') as f:
                sheetName = os.path.basename(path)[:-4]
                startRow = 3
                readCurrentRow = 1
                rowCount = 1
                tableRow = 1
                self.currentCol = 2
                self.row = 3
                self.col = 2

                # Reading csv file row by row.
                # Inputs: None.
                for row in f:

                    valList = row.split(',')

                    # Adding describing values to an object list for titling purposes.
                    # Table column descriptors.
                    if valList[2] not in self.descList:
                        self.descList.append(valList[2])

                    # Reading variable to determine if it is the first row of csv file.
                    # Creates a sheet in the workbook for describing comments of data.
                    if firstRun == 0:
                        numOfRows, numOfCols = self.testSpecificFunc[test](0, med, valList, sheetName, None)
                        firstRun = 1
                        self.wb.save(filename=f"{self.destFolder}/{self.finalReport}")
                        continue

                    # Reading variable to determine if it is the second row of csv file.
                    # Creates a list of names to title each table.
                    elif firstRun == 1:
                        self.testSpecificFunc[test](line=1, med=None, valList=valList, sheetName=None, numOfCols=numOfCols)
                        firstRun = 2
                        continue

                    # Reading variable to determine it is not the first or second row of csv file.
                    else:
                        # Opening a sheet as the title of the csv, continuing to write for that sheet for the remainder
                        # of the csv file reading.
                        if f'{sheetName}' not in self.wb.sheetnames:
                            ws = self.wb.create_sheet(title=f'{sheetName}')
                        else:
                            ws = self.wb[f'{sheetName}']

                        self.currentRow = readCurrentRow + 3
                        # Placing load steps into cells: only placed into column=1.
                        for table in range(numOfCols):
                            ws.cell(row=self.currentRow, column=1).value = float(valList[3])
                            ws.cell(row=self.currentRow, column=1).border = Border(top=med, bottom=med, left=med,
                                                                                   right=med)
                            ws.cell(row=self.currentRow, column=1).alignment = Alignment(horizontal='center')
                            self.currentRow = self.currentRow + (numOfRows + 3)

                        table = 0
                        # Placing data into cells based upon rows in a single column.
                        for value in valList[4:]:
                            self.currentRow = (table * (numOfRows + 3)) + tableRow + 3
                            table += 1
                            ws.cell(row=self.currentRow, column=self.currentCol).value = float(value)
                            ws.cell(row=self.currentRow, column=self.currentCol).border = Border(top=med, bottom=med,
                                                                                                 left=med, right=med)
                            ws.cell(row=self.currentRow, column=self.currentCol).alignment = Alignment(
                                horizontal='center')

                    # Increasing column dimension if the rows have filled in current column.
                    if rowCount == numOfRows:
                        self.currentCol += 1
                        readCurrentRow = 0
                        rowCount = 0
                        tableRow = 0

                    readCurrentRow += 1
                    rowCount += 1
                    tableRow += 1

                # Insert all column titles.
                chartPlacement = 2
                chartTitle = 1
                for i in range(1, len(self.tableList)):
                    ws.cell(row=self.row - 1, column=1).value = self.tableList[i]  # adding column titles
                    ws.cell(row=self.row - 1, column=1).fill = PatternFill(start_color="63d5a0", end_color="63d5a0",
                                                                           fill_type="solid")
                    ws.cell(row=self.row - 1, column=1).alignment = Alignment(horizontal='center')
                    ws.cell(row=self.row, column=1).value = self.tableList[0]  # adding load title
                    ws.cell(row=self.row, column=1).border = Border(top=med, bottom=med, left=med, right=med)
                    ws.cell(row=self.row, column=1).fill = PatternFill(start_color="EEE8AA", end_color="EEE8AA",
                                                                       fill_type="solid")
                    ws.cell(row=self.row, column=1).alignment = Alignment(horizontal='center')

                    #region Chart
                    chart = ScatterChart()
                    xvalues = Reference(ws, min_col=1, min_row=startRow+1, max_row=numOfRows + startRow)

                    # Create different 'y' values for the graph based upon each column in the current table.
                    for val in range(2, len(self.descList)):
                        values = Reference(ws, min_col=val, min_row=startRow, max_row=numOfRows + startRow)
                        series = Series(values, xvalues, title_from_data=True)
                        chart.series.append(series)
                    chart.title = self.tableList[chartTitle]  # Titling chart based upon table title.
                    chart.y_axis.title = self.tableList[chartTitle]
                    chart.x_axis.title = 'Load (A)'

                    chart.width = 14
                    chart.height = 8
                    num = util.get_column_letter(len(self.descList) + 1)
                    ws.add_chart(chart, f"{num}{chartPlacement}")
                    chartPlacement += numOfRows + 3
                    startRow += numOfRows + 3
                    chartTitle += 1
                    #endregion

                    # Insert table descriptor for each table.
                    for ind in range(len(self.descList)):
                        # For first and second values in descriptor list, do not write.
                        # These values are for data inputs regarding organization.
                        if ind == 0:
                            continue
                        if ind == 1:
                            ws.column_dimensions[f'{util.get_column_letter(ind)}'].width = 40
                            continue
                        ws.column_dimensions[f'{util.get_column_letter(ind)}'].width = 40
                        ws.cell(row=self.row, column=self.col).value = self.descList[ind]
                        ws.cell(row=self.row, column=self.col).border = Border(top=med, bottom=med, left=med, right=med)
                        ws.cell(row=self.row, column=self.col).fill = PatternFill(start_color="EEE8AA",
                                                                                  end_color="EEE8AA", fill_type="solid")
                        ws.cell(row=self.row, column=self.col).alignment = Alignment(horizontal='center')
                        self.col = self.col + 1
                    self.row = self.row + numOfRows + 3
                    self.col = 2

                # Remove the autogenerated worksheet titled "Sheet".
                if 'Sheet' in self.wb.sheetnames:
                    self.wb.remove(self.wb['Sheet'])

            self.wb.save(filename=f"{self.destFolder}/{self.finalReport}")
        self.wb.close()


if __name__ == '__main__':
    rs = RunReportGen()
    rs._get_file_paths()
    rs.destFolder = '.'
    rs.generic_report_gen('eff')
